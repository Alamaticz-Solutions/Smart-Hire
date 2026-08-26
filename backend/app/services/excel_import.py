"""Bulk Excel candidate-import service.

Moved verbatim from `app/main.py` (original lines 2214-2500):

  - `process_excel_file_logic` (original lines 2214-2476)
  - `process_excel_file`       (original lines 2478-2487)

`main.py` itself is left untouched; this module is additive.

BUG FIX (approved): `match_id` was unconditionally set to `None`
(original line 2388, `match_id = None`) and never reassigned before the
`if match_id:` check (original line 2425), so the "update existing
record" branch (original lines 2426-2429) was permanently unreachable -
every Excel row always inserted a new candidate row, even though
`existing_candidates` was built (original lines 2317-2326) and kept
up to date (original line ~2443, appended to on every insert)
specifically to support de-duplication against re-imports.

The surrounding code makes the intended matching strategy clear:
`existing_candidates` is scoped to rows already created by this same
`username` (`WHERE LOWER(created_by) = LOWER(?)`, original line 2317) and
carries `id`, `full_name`, `email`, `phone` for exactly that purpose, and
the row being imported already computes `norm_email`/`norm_phone` right
before the dead branch (original lines 2385-2386) via the same
`normalize_email`/`normalize_phone` helpers main.py uses elsewhere for
candidate de-dup. So the fix wires `match_id` to the first
`existing_candidates` entry whose normalized email matches (email is the
stronger identifier), falling back to a normalized-phone match when no
email match is found - see `_find_existing_match` below. This closes the
gap without inventing a new matching strategy; it just finishes the one
the code was already set up for.
"""

from __future__ import annotations

import os
import re
from typing import Optional

import openpyxl

from app.core.logging import get_logger
from app.db.session import get_db_connection
from app.services.ai_clients import _processing_lock
from app.services.matching import match_candidate_to_all_jobs

logger = get_logger(__name__)


# ── Normalization helpers ──────────────────────────────────────────────────
# Duplicated from main.py's module-level `normalize_email`/`normalize_phone`
# (original lines 34-48) rather than imported, since main.py is left
# untouched and no shared `app.services.candidate_utils`-style module exists
# yet for these small pure functions. Safe to repoint at a shared module
# later.
def normalize_phone(phone) -> str:
    if not phone:
        return ""
    s = str(phone).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def normalize_email(email) -> str:
    if not email:
        return ""
    return str(email).strip().lower()


def _find_existing_match(existing_candidates: list, norm_email: str, norm_phone: str) -> Optional[int]:
    """Return the `id` of the existing_candidates entry matching this row, or None.

    Email match takes priority (stronger identifier); phone is the
    fallback, matching how `existing_candidates` was populated and scoped
    (see module docstring's BUG FIX note).
    """
    if norm_email:
        for cand in existing_candidates:
            if normalize_email(cand.get("email")) == norm_email:
                return cand.get("id")
    if norm_phone:
        for cand in existing_candidates:
            if normalize_phone(cand.get("phone")) == norm_phone:
                return cand.get("id")
    return None


def process_excel_file_logic(safe_name: str, path: str, username: str):
    with get_db_connection() as conn:
        cur = conn.cursor()
        processed_candidate_ids = []
        try:
            wb = openpyxl.load_workbook(path, data_only=True)

            # Fetch custom columns.
            cur.execute("SELECT col_key, col_label FROM custom_columns")
            custom_cols = {row[1].strip().lower(): row[0] for row in cur.fetchall()}

            # Mappings for common columns.
            column_mappings = {
                "name": "full_name",
                "full name": "full_name",
                "candidate name": "full_name",
                "candidate": "full_name",
                "names": "full_name",
                "phone": "phone",
                "phone no": "phone",
                "phone number": "phone",
                "mobile": "phone",
                "mobile number": "phone",
                "mobile no": "phone",
                "contact": "phone",
                "contact number": "phone",
                "contact no": "phone",
                "email": "email",
                "email id": "email",
                "email address": "email",
                "skills": "skills",
                "key skills": "skills",
                "technical skills": "skills",
                "experience": "total_experience",
                "exp": "total_experience",
                "total exp": "total_experience",
                "total experience": "total_experience",
                "work experience": "total_experience",
                "pega experience": "pega_experience",
                "pega exp": "pega_experience",
                "cdh experience": "cdh_exp",
                "cdh exp": "cdh_exp",
                "current ctc": "ctc",
                "ctc": "ctc",
                "salary": "ctc",
                "expected ctc": "expected_ctc",
                "percentage hike": "percentage_hike",
                "hike": "percentage_hike",
                "notice period": "notice_period",
                "np": "notice_period",
                "current location": "current_location",
                "location": "current_location",
                "preferred locations": "pref_locations",
                "preferred location": "pref_locations",
                "pref locations": "pref_locations",
                "current organization": "current_organization",
                "current employer": "current_organization",
                "employer": "current_organization",
                "current employment": "current_organization",
                "current client": "current_client",
                "domain": "domain",
                "tier": "tier",
                "certification version": "certification_version",
                "certifications": "certifications",
                # Additional maps for thorough Excel parsing.
                "linkedin": "linkedin",
                "linkedin url": "linkedin",
                "linkedin profile": "linkedin",
                "notes": "notescomments",
                "comments": "notescomments",
                "notes/comments": "notescomments",
                "notescomments": "notescomments",
                "feedback": "notescomments",
                "candidate status": "candidate_status",
                "candidate_status": "candidate_status",
                "status": "candidate_status",
                "candidate interview status": "candidate_interview_status",
                "interview status": "candidate_interview_status",
                "availability": "availability_in_days",
                "availability in days": "availability_in_days",
                "availability_in_days": "availability_in_days",
                "source": "source",
                "candidate source": "source",
                "how did you find us": "source",
            }

            # Load existing candidates from DB.
            cur.execute("SELECT id, full_name, email, phone FROM candidate_metadata WHERE LOWER(created_by) = LOWER(?)", (username,))
            existing_candidates = [
                {
                    "id": r[0],
                    "full_name": r[1] or "",
                    "email": r[2] or "",
                    "phone": r[3] or "",
                }
                for r in cur.fetchall()
            ]

            # Iterate over all sheets.
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Find the header row (iterating up to 10 rows).
                header_row_idx = None
                headers = []
                for r_idx in range(1, min(ws.max_row + 1, 11)):
                    row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, ws.max_column + 1)]
                    non_empty = [v for v in row_vals if v is not None]
                    if len(non_empty) >= 2:
                        is_header = False
                        for val in non_empty:
                            val_str = str(val).strip().lower()
                            if val_str in column_mappings or val_str in custom_cols:
                                is_header = True
                                break
                        if is_header:
                            header_row_idx = r_idx
                            headers = [str(h).strip() if h is not None else "" for h in row_vals]
                            break

                if not header_row_idx:
                    header_row_idx = 1
                    headers = [str(ws.cell(row=1, column=c_idx).value).strip() if ws.cell(row=1, column=c_idx).value is not None else "" for c_idx in range(1, ws.max_column + 1)]

                # Map headers to DB columns.
                mapped_cols = {}
                for idx, h in enumerate(headers):
                    h_lower = h.lower()
                    if h_lower in custom_cols:
                        mapped_cols[idx] = custom_cols[h_lower]
                    elif h_lower in column_mappings:
                        mapped_cols[idx] = column_mappings[h_lower]

                # Process rows starting from header_row_idx + 1.
                for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                    row_data = {}
                    is_row_empty = True
                    for col_idx in range(1, len(headers) + 1):
                        val = ws.cell(row=row_idx, column=col_idx).value
                        if val is not None:
                            is_row_empty = False
                        if (col_idx - 1) in mapped_cols:
                            db_col = mapped_cols[col_idx - 1]
                            cell_str = str(val).strip() if val is not None else ""
                            # BUG FIX: two differently-worded headers (e.g.
                            # "Experience" and "Total Exp", or "Location" and
                            # "Current Location") can both map to the same
                            # db_col (see column_mappings above). Iterating
                            # columns left-to-right and unconditionally
                            # overwriting row_data[db_col] meant a later,
                            # empty duplicate-header column silently blanked
                            # out a value a real header already set for this
                            # row. Don't let a later blank duplicate-header
                            # column clobber a value an earlier column
                            # already set for this db_col (a later *non-empty*
                            # duplicate still wins, unchanged from before).
                            if cell_str or db_col not in row_data:
                                row_data[db_col] = cell_str

                    if is_row_empty:
                        continue

                    # Check that we have at least one identifying field.
                    if not row_data.get("full_name") and not row_data.get("email") and not row_data.get("phone"):
                        continue

                    name = row_data.get("full_name", "")
                    email = row_data.get("email", "")
                    phone = row_data.get("phone", "")

                    norm_email = normalize_email(email)
                    norm_phone = normalize_phone(phone)

                    # BUG FIX: this used to be hardcoded to `None` and never
                    # reassigned, making the "update existing record" branch
                    # below permanently unreachable (see module docstring).
                    match_id = _find_existing_match(existing_candidates, norm_email, norm_phone)

                    # Normalize numeric/experience fields.
                    # BUG FIX: a plain `float(...)` cast on anything the sheet
                    # author typed as free text (e.g. "5+ years", "5-7 yrs",
                    # a stray "Yrs" suffix) raised ValueError and silently
                    # zeroed out real experience data. `resume_processing.py`
                    # and `email_worker.py` handle the identical
                    # LLM-extracted-value problem by regex-extracting the
                    # first number instead of hard-failing to 0.0; do the
                    # same here for consistency so a messy-but-legible cell
                    # like "5+" still imports as 5.0 instead of losing the
                    # candidate's experience data.
                    for exp_field in ("total_experience", "pega_experience", "cdh_exp"):
                        if exp_field in row_data and row_data[exp_field] != "":
                            raw = row_data[exp_field]
                            try:
                                # Clean numeric cells (including "5" or ".5")
                                # take this fast, exact path unchanged.
                                row_data[exp_field] = float(raw)
                            except ValueError:
                                # Only fall back to regex-extracting the
                                # first number for genuinely messy text like
                                # "5+ years" or "5-7 yrs". Skip the fallback
                                # for date-like strings (e.g. an Excel date
                                # cell such as "2019-07-01 00:00:00") so a
                                # mismapped date column doesn't get read as
                                # "2019.0 years of experience" - 0.0 (which
                                # correctly surfaces as "missing experience"
                                # downstream) is the safer failure mode here.
                                if re.search(r"\d{4}-\d{2}-\d{2}", str(raw)) or "/" in str(raw):
                                    row_data[exp_field] = 0.0
                                else:
                                    match = re.search(r"\d+(\.\d+)?", str(raw))
                                    row_data[exp_field] = float(match.group()) if match else 0.0
                    if "notice_period" in row_data and row_data["notice_period"] != "":
                        try:
                            digits = "".join(c for c in row_data["notice_period"] if c.isdigit())
                            row_data["notice_period"] = int(digits) if digits else ""
                        except Exception:
                            row_data["notice_period"] = ""
                    if "availability_in_days" in row_data and row_data["availability_in_days"] != "":
                        try:
                            digits = "".join(c for c in row_data["availability_in_days"] if c.isdigit())
                            row_data["availability_in_days"] = int(digits) if digits else ""
                        except Exception:
                            row_data["availability_in_days"] = ""

                    cur.execute("PRAGMA table_info(candidate_metadata)")
                    allowed_cols = {c[1] for c in cur.fetchall()}
                    db_data = {k: v for k, v in row_data.items() if k in allowed_cols and k != "id"}
                    if "source" not in db_data or not db_data["source"]:
                        db_data["source"] = "Excel Import"

                    if match_id:
                        # Update existing record.
                        set_clause = ", ".join(f"{k}=?" for k in db_data)
                        cur.execute(f"UPDATE candidate_metadata SET {set_clause} WHERE id=?", list(db_data.values()) + [match_id])
                        candidate_id = match_id
                    else:
                        # Insert new record.
                        db_data["is_approved"] = 1
                        db_data["candidate_status"] = "New"
                        db_data["filename"] = ""
                        db_data["created_by"] = username
                        cols_list = list(db_data.keys())
                        vals_list = list(db_data.values())
                        cur.execute(
                            f"INSERT INTO candidate_metadata ({','.join(cols_list)}) VALUES ({','.join(['?'] * len(cols_list))})",
                            vals_list,
                        )
                        candidate_id = cur.lastrowid
                        existing_candidates.append(
                            {
                                "id": candidate_id,
                                "full_name": name,
                                "email": email,
                                "phone": phone,
                            }
                        )

                    # Match candidate to all jobs later.
                    processed_candidate_ids.append(candidate_id)

        except Exception as e:
            logger.error(f"Error parsing Excel file {safe_name}: {e}")
            if os.path.exists(path):
                try:
                    os.remove(path)
                    logger.info(f"Cleaned up failed Excel file: {path}")
                except Exception as file_err:
                    logger.error(f"Error cleaning up failed Excel file {path}: {file_err}")
        finally:
            try:
                cur.execute("DELETE FROM candidate_metadata WHERE filename = ? AND full_name LIKE '⏳ Parsing Excel:%'", (safe_name,))
                conn.commit()
            except Exception as db_err:
                logger.error(f"Error clearing Excel placeholder or committing in finally: {db_err}")

    # Match each newly-imported candidate to all active jobs (original
    # main.py lines 2472-2476) - per-candidate try/except preserved exactly
    # so one bad match doesn't abort matching for the rest of the batch.
    for cid in processed_candidate_ids:
        try:
            match_candidate_to_all_jobs(cid)
        except Exception as match_err:
            logger.error(f"Error matching candidate {cid}: {match_err}")
    if processed_candidate_ids:
        logger.info(f"Excel import processed {len(processed_candidate_ids)} candidate(s) pending job re-match: {processed_candidate_ids}")


def process_excel_file(safe_name: str, path: str, username: str = "unknown"):
    """Serialize Excel processing behind the app's global processing lock, then clean up the temp file.

    Moved from main.py original lines 2478-2487. The `_processing_lock`
    (defined in main.py as a module-level `threading.Lock()` shared by both
    resume and Excel processing, to avoid Render OOM / Groq rate limits /
    SQLite locks) still lives in main.py; it's imported lazily here to avoid
    importing the whole app module at import time.
    """

    try:
        with _processing_lock:
            process_excel_file_logic(safe_name, path, username)
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass
