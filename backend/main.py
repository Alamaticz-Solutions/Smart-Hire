import os
import json
import sqlite3
import shutil
import re
import gc
import threading
from typing import Optional
from datetime import datetime
import openpyxl

def normalize_phone(phone) -> str:
    if not phone:
        return ""
    s = str(phone).strip()
    if s.endswith('.0'):
        s = s[:-2]
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) >= 10:
        return digits[-10:]
    return digits

def normalize_email(email) -> str:
    if not email:
        return ""
    return str(email).strip().lower()

def get_base_email(email: str) -> str:
    if not email:
        return ""
    email_str = str(email).strip().lower()
    if "@" in email_str:
        parts = email_str.split("@")
        local_part = parts[0].split("+")[0]
        return f"{local_part}@{parts[1]}"
    return email_str

def is_similar_name(n1: str, n2: str) -> bool:
    if not n1 or not n2:
        return False
    # Clean non-alphanumeric chars and split into lowercase words
    w1 = set(re.findall(r'[a-z0-9]+', n1.lower()))
    w2 = set(re.findall(r'[a-z0-9]+', n2.lower()))
    if not w1 or not w2:
        return False
    # If they are exactly equal
    if n1.strip().lower() == n2.strip().lower():
        return True
    
    # Calculate intersection
    intersection = w1.intersection(w2)
    if not intersection:
        return False
        
    # We want at least two common words to match (e.g. "durga" and "dheeraj")
    # unless one of them is only a single word, in which case we check if it is part of the other
    shorter_len = min(len(w1), len(w2))
    if shorter_len == 1:
        word = list(w1 if len(w1) == 1 else w2)[0]
        if len(word) >= 4:
            return word in w1 or word in w2
        return False
    
    return len(intersection) >= 2

def phones_match(p1: str, p2: str) -> bool:
    # Get only the digits of both phone numbers
    d1 = "".join(c for c in str(p1) if c.isdigit())
    d2 = "".join(c for c in str(p2) if c.isdigit())
    if not d1 or not d2:
        return False
    # If they are exactly equal
    if d1 == d2:
        return True
    # If the last 10 digits are equal
    if len(d1) >= 10 and len(d2) >= 10 and d1[-10:] == d2[-10:]:
        return True
    # Check if the shorter one is contained in the longer one, as long as it's at least 9 digits
    shorter, longer = (d1, d2) if len(d1) < len(d2) else (d2, d1)
    if len(shorter) >= 9 and shorter in longer:
        return True
    return False

from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Request, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, FileResponse
from starlette.exceptions import HTTPException as StarletteHTTPException
from pydantic import BaseModel
from dotenv import load_dotenv
import fitz
from langchain_core.documents import Document

class SafePyMuPDFLoader:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def load(self):
        docs = []
        try:
            doc = fitz.open(self.file_path) 
            for page_num, page in enumerate(doc):
                text = page.get_text()
                metadata = {
                    "source": self.file_path,
                    "page": page_num,
                }
                docs.append(Document(page_content=text, metadata=metadata))
            doc.close()
        except Exception as e:
            print(f"Error loading PDF {self.file_path} with fitz: {e}")
            raise e
        return docs

# LangChain / AI
from langchain_community.document_loaders import Docx2txtLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_chroma import Chroma
from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import PromptTemplate

# ── Config ────────────────────────────────────────────────────────────────────
load_dotenv()
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(BASE_DIR)

# Support persistent volume directory /data (e.g. on Render)
DATA_DIR = "/data" if os.path.exists("/data") and os.access("/data", os.W_OK) else BASE_DIR

CHROMA_PATH  = os.path.join(DATA_DIR, "chroma_db")
UPLOAD_DIR   = os.path.join(DATA_DIR, "static")
STATS_DB     = os.getenv("STATS_DB_PATH", os.path.join(DATA_DIR, "stats.db"))
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(CHROMA_PATH, exist_ok=True)
os.makedirs(os.path.dirname(STATS_DB), exist_ok=True)

# ── FastAPI App ────────────────────────────────────────────────────────────────
app = FastAPI(title="Hire AI API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve uploaded resumes
app.mount("/static", StaticFiles(directory=UPLOAD_DIR), name="static")

# ── Models (loaded once) ───────────────────────────────────────────────────────
_embeddings = None
_llm        = None
_active_groq_key = None
_models_loading = False
_model_lock = threading.Lock()
_processing_lock = threading.Lock()

def get_models():
    global _embeddings, _llm, _models_loading, _active_groq_key
    with _model_lock:
        _models_loading = True
        if _embeddings is None:
            try:
                _embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
            except Exception as e:
                print(f"Warning: Failed to load HuggingFaceEmbeddings: {e}")
                _embeddings = None
        
        # Load env variables dynamically in case the key has changed in the file
        try:
            from dotenv import load_dotenv
            load_dotenv(override=True)
        except Exception:
            pass
        current_key = os.getenv("GROQ_API_KEY", "")
        
        if _llm is None or _active_groq_key != current_key:
            _llm = ChatGroq(temperature=0.1, model_name="llama-3.1-8b-instant", groq_api_key=current_key)
            _active_groq_key = current_key
            
        _models_loading = False
    return _embeddings, _llm

@app.on_event("startup")
def load_models_in_background():
    threading.Thread(target=get_models, daemon=True).start()
    threading.Thread(target=poll_emails_and_process, daemon=True).start()

# ── DB Helpers ─────────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur  = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS candidate_metadata (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            filename             TEXT,
            full_name            TEXT,
            candidate_status     TEXT DEFAULT 'New',
            total_experience     REAL DEFAULT 0.0,
            pega_experience      REAL DEFAULT 0.0,
            skills               TEXT,
            certifications       TEXT,
            ctc                  TEXT,
            notice_period        TEXT,
            current_organization TEXT,
            email                TEXT,
            phone                TEXT,
            linkedin             TEXT,
            created_by           TEXT DEFAULT 'admin',
            timestamp            DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS custom_columns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            col_key TEXT UNIQUE,
            col_label TEXT,
            description TEXT
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT NOT NULL,
            client_name TEXT,
            contact_name TEXT,
            client_phone TEXT,
            account_manager TEXT,
            assigned_recruiter TEXT,
            target_date TEXT,
            job_type TEXT,
            job_status TEXT,
            work_experience TEXT,
            industry TEXT,
            salary TEXT,
            required_skills TEXT,
            created_by TEXT DEFAULT 'admin',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS job_candidates (
            job_id INTEGER,
            candidate_id INTEGER,
            ai_reason TEXT,
            status TEXT DEFAULT 'matched',
            PRIMARY KEY (job_id, candidate_id),
            FOREIGN KEY (job_id) REFERENCES jobs(id),
            FOREIGN KEY (candidate_id) REFERENCES candidate_metadata(id)
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            is_hr INTEGER DEFAULT 0,
            is_admin INTEGER DEFAULT 0,
            is_external INTEGER DEFAULT 0
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS change_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            action_type TEXT NOT NULL,
            target_id TEXT,
            payload TEXT,
            description TEXT,
            status TEXT DEFAULT 'pending',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS activity_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            action TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS team_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS job_shares (
            job_id INTEGER,
            username TEXT,
            PRIMARY KEY (job_id, username),
            FOREIGN KEY (job_id) REFERENCES jobs(id),
            FOREIGN KEY (username) REFERENCES users(username)
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS masked_keywords (
            keyword TEXT PRIMARY KEY
        )
    ''')
    cur.execute("SELECT COUNT(*) FROM masked_keywords")
    if cur.fetchone()[0] == 0:
        for kw in ["CSSA", "LSA"]:
            cur.execute("INSERT OR IGNORE INTO masked_keywords (keyword) VALUES (?)", (kw,))

    # Migration: add missing columns to candidate_metadata
    cur.execute("PRAGMA table_info(candidate_metadata)")
    existing = [c[1] for c in cur.fetchall()]
    new_cols = {
        'candidate_status': "TEXT DEFAULT 'New'",
        'source': "TEXT DEFAULT 'Resume Upload'",
        'cdh_exp': 'REAL DEFAULT 0.0',
        'expected_ctc': 'TEXT',
        'percentage_hike': 'TEXT',
        'candidate_interview_status': 'TEXT',
        'availability_in_days': 'INTEGER',
        'current_location': 'TEXT',
        'pref_locations': 'TEXT',
        'current_client': 'TEXT',
        'domain': 'TEXT',
        'tier': 'TEXT',
        'certification_version': 'TEXT',
        'current_organization': 'TEXT',
        'email': 'TEXT',
        'phone': 'TEXT',
        'linkedin': 'TEXT',
        'email_message': 'TEXT',
        'formatted_json': 'TEXT',
        'sender_email': 'TEXT',
        'is_qualified': 'INTEGER DEFAULT 0',
        'is_approved': 'INTEGER DEFAULT 1',
        'created_by': "TEXT DEFAULT 'admin'"
    }
    for col, dtype in new_cols.items():
        if col not in existing:
            cur.execute(f"ALTER TABLE candidate_metadata ADD COLUMN {col} {dtype}")

    # Migration: add missing columns to jobs
    cur.execute("PRAGMA table_info(jobs)")
    existing_jobs = [c[1] for c in cur.fetchall()]
    new_jobs_cols = {
        'client_name': 'TEXT',
        'contact_name': 'TEXT',
        'client_phone': 'TEXT',
        'account_manager': 'TEXT',
        'assigned_recruiter': 'TEXT',
        'target_date': 'TEXT',
        'job_type': 'TEXT',
        'job_status': 'TEXT',
        'work_experience': 'TEXT',
        'industry': 'TEXT',
        'salary': 'TEXT',
        'required_skills': 'TEXT',
        'created_by': "TEXT DEFAULT 'admin'"
    }
    for col, dtype in new_jobs_cols.items():
        if col not in existing_jobs:
            cur.execute(f"ALTER TABLE jobs ADD COLUMN {col} {dtype}")
            
    # Migration: add missing columns to users
    cur.execute("PRAGMA table_info(users)")
    existing_users_cols = [c[1] for c in cur.fetchall()]
    if 'is_hr' not in existing_users_cols:
        cur.execute("ALTER TABLE users ADD COLUMN is_hr INTEGER DEFAULT 0")
    if 'is_admin' not in existing_users_cols:
        cur.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER DEFAULT 0")
    if 'is_external' not in existing_users_cols:
        cur.execute("ALTER TABLE users ADD COLUMN is_external INTEGER DEFAULT 0")
    if 'hidden_fields' not in existing_users_cols:
        cur.execute("ALTER TABLE users ADD COLUMN hidden_fields TEXT DEFAULT ''")
    if 'email' not in existing_users_cols:
        cur.execute("ALTER TABLE users ADD COLUMN email TEXT")
    if 'is_approved' not in existing_users_cols:
        cur.execute("ALTER TABLE users ADD COLUMN is_approved INTEGER DEFAULT 0")

    # Seed default users if empty (do NOT wipe the table to preserve registered users!)
    cur.execute("SELECT COUNT(*) FROM users WHERE LOWER(username) = 'admin'")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO users (full_name, username, password, role, is_hr, is_admin, email) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    ("System Admin", "admin", "admin", "admin", 1, 1, "admin@gmail.com"))
    else:
        # Guarantee admin password is 'admin' as requested by the user
        cur.execute("UPDATE users SET password = ?, role = 'admin', is_hr = 1, is_admin = 1 WHERE LOWER(username) = 'admin'", ("admin",))

    cur.execute("SELECT COUNT(*) FROM users WHERE LOWER(username) = 'user'")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO users (full_name, username, password, role, is_hr, is_admin, email) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    ("Test User", "user", "user", "user", 0, 0, "user@gmail.com"))

    # Seed the 4 recruiters in user management if they do not exist
    for m in ["Boopathi", "Praveen", "Harish", "Sabari"]:
        uname = m.lower()
        role = "admin" if uname == "sabari" else "user"
        is_admin = 1 if uname == "sabari" else 0
        is_hr = 1 if uname == "sabari" else 0
        cur.execute("SELECT COUNT(*) FROM users WHERE LOWER(username) = ?", (uname,))
        if cur.fetchone()[0] == 0:
            cur.execute("INSERT INTO users (full_name, username, password, role, is_hr, is_admin, email) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (m, uname, uname, role, is_hr, is_admin, f"{uname}@gmail.com"))
        else:
            cur.execute("UPDATE users SET is_hr = ?, is_admin = ?, role = ? WHERE LOWER(username) = ?", (is_hr, is_admin, role, uname))
        
    cur.execute("SELECT COUNT(*) FROM team_members")
    if cur.fetchone()[0] == 0:
        for m in ["Boopathi", "Praveen", "Harish", "Sabari"]:
            cur.execute("INSERT OR IGNORE INTO team_members (name) VALUES (?)", (m,))
        
    # Pre-approve seeded/default users
    cur.execute("UPDATE users SET is_approved = 1 WHERE LOWER(username) IN ('admin', 'user', 'boopathi', 'praveen', 'harish', 'sabari')")

    # Fix emails of existing seeded/default users and Somasekhar9 if they are NULL/empty
    for uname in ["admin", "user", "boopathi", "praveen", "harish", "sabari", "somasekhar9"]:
        cur.execute("UPDATE users SET email = ? WHERE LOWER(username) = ? AND (email IS NULL OR email = '')", (f"{uname}@gmail.com", uname))

    # Create integrations_settings table
    cur.execute('''
    CREATE TABLE IF NOT EXISTS integrations_settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email_enabled INTEGER DEFAULT 0,
        imap_host TEXT DEFAULT 'imap.gmail.com',
        imap_port INTEGER DEFAULT 993,
        smtp_host TEXT DEFAULT 'smtp.gmail.com',
        smtp_port INTEGER DEFAULT 587,
        email_user TEXT,
        email_pass TEXT,
        keywords TEXT DEFAULT 'resume,alamaticz,solution,job',
        drive_enabled INTEGER DEFAULT 0
    )
    ''')

    # Seed integrations_settings if empty, loading from environment variables if present
    cur.execute("SELECT COUNT(*) FROM integrations_settings")
    if cur.fetchone()[0] == 0:
        env_user = os.getenv("SMTP_SENDER", "")
        env_pass = os.getenv("SMTP_PASSWORD", "")
        env_imap = os.getenv("IMAP_HOST", "imap.gmail.com")
        env_smtp = os.getenv("SMTP_HOST", "smtp.gmail.com")
        email_enabled = 1 if env_user and env_pass else 0
        cur.execute("""
        INSERT INTO integrations_settings (email_enabled, imap_host, imap_port, smtp_host, smtp_port, email_user, email_pass, keywords, drive_enabled)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (email_enabled, env_imap, 993, env_smtp, 587, env_user, env_pass, "resume,alamaticz,solution,job", 0))

    # Create processed_emails table
    cur.execute('''
    CREATE TABLE IF NOT EXISTS processed_emails (
        msg_uid TEXT PRIMARY KEY,
        processed_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    conn.commit()
    conn.close()

def get_candidates_list(username: Optional[str] = None, role: str = "user") -> list:
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    try:
        is_hr_or_admin = False
        if username:
            is_hr_or_admin = is_admin_or_hr(username)
        if role == "admin" or is_hr_or_admin or not username:
            cur.execute("SELECT * FROM candidate_metadata ORDER BY timestamp DESC")
        else:
            cur.execute("SELECT * FROM candidate_metadata WHERE LOWER(created_by) = LOWER(?) ORDER BY timestamp DESC", (username,))
        rows = [dict(r) for r in cur.fetchall()]
    except Exception:
        rows = []
    conn.close()
    return rows

def get_masked_keywords() -> list:
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    try:
        cur.execute("SELECT keyword FROM masked_keywords")
        keywords = [row[0] for row in cur.fetchall()]
    except Exception:
        keywords = []
    conn.close()
    return keywords

def mask_text_with_keywords(text: str, keywords: list) -> str:
    if not text or not keywords:
        return text
    result = str(text)
    for kw in keywords:
        if not kw.strip():
            continue
        pattern = re.compile(re.escape(kw), re.IGNORECASE)
        result = pattern.sub("****", result)
    return result

def mask_candidate_record(candidate: dict, keywords: list) -> dict:
    masked = {}
    for k, v in candidate.items():
        if isinstance(v, str):
            masked[k] = mask_text_with_keywords(v, keywords)
        else:
            masked[k] = v
    return masked


def log_candidate(data: dict):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur  = conn.cursor()
    
    cur.execute("PRAGMA table_info(candidate_metadata)")
    existing_cols = {c[1] for c in cur.fetchall()}
    
    # Filter data to only valid columns
    cols = [c for c in data.keys() if c in existing_cols and c != 'id']
    vals = [str(data.get(c, '')) if data.get(c) is not None else '' for c in cols]
    
    if data.get('filename'):
        cur.execute("DELETE FROM candidate_metadata WHERE filename = ?", (data.get('filename'),))
    
    new_id = None
    if cols:
        cur.execute(
            f"INSERT INTO candidate_metadata ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})",
            vals
        )
        new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return new_id

init_db()

def log_activity_db(username: str, action: str):
    if not username:
        username = "unknown"
    try:
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        cur = conn.cursor()
        cur.execute("INSERT INTO activity_logs (username, action) VALUES (?, ?)", (username, action))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error logging activity: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

# ── Health ─────────────────────────────────────────────────────────────────────
@app.get("/api/health")
def health():
    return {"status": "ok"}

class ActivityCreate(BaseModel):
    username: str
    action: str

@app.get("/api/activity")
def get_activity_logs():
    try:
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM activity_logs ORDER BY timestamp DESC")
        logs = [dict(row) for row in cur.fetchall()]
        conn.close()
        return logs
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/activity")
def create_activity_log(req: ActivityCreate):
    log_activity_db(req.username, req.action)
    return {"status": "logged"}

@app.delete("/api/activity")
def clear_activity_logs(request: Request):
    username = request.headers.get("x-user-username")
    try:
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        cur = conn.cursor()
        cur.execute("DELETE FROM activity_logs")
        conn.commit()
        conn.close()
        log_activity_db(username or "unknown", "cleared the activity feed")
        return {"status": "cleared"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class TeamMemberCreate(BaseModel):
    name: str

@app.get("/api/team-members")
def list_team_members():
    try:
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM team_members ORDER BY name ASC")
        members = [dict(row) for row in cur.fetchall()]
        conn.close()
        return members
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/team-members")
def create_team_member(req: TeamMemberCreate, request: Request):
    username = request.headers.get("x-user-username") or "admin"
    name = req.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name cannot be empty")
    try:
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        cur = conn.cursor()
        try:
            cur.execute("INSERT INTO team_members (name) VALUES (?)", (name,))
            conn.commit()
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail=f"Team member '{name}' already exists")
        finally:
            conn.close()
        log_activity_db(username, f"added '{name}' to the recruiter persona matrix")
        return {"status": "added", "name": name}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/team-members/{member_id}")
def delete_team_member(member_id: int, request: Request):
    username = request.headers.get("x-user-username") or "admin"
    try:
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        cur = conn.cursor()
        
        # Get member name first to log it
        cur.execute("SELECT name FROM team_members WHERE id = ?", (member_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="Team member not found")
        member_name = row[0]
        
        cur.execute("DELETE FROM team_members WHERE id = ?", (member_id,))
        conn.commit()
        conn.close()
        
        log_activity_db(username, f"removed '{member_name}' from the recruiter persona matrix")
        return {"status": "deleted"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def is_user_approved(username: Optional[str]) -> bool:
    if not username:
        return False
    # Seeded/default users are always approved
    if username.lower() in ("admin", "user", "boopathi", "praveen", "harish", "sabari"):
        return True
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    try:
        cur.execute("SELECT is_approved FROM users WHERE LOWER(username) = LOWER(?)", (username,))
        row = cur.fetchone()
        if row and row[0] == 1:
            return True
    except Exception:
        pass
    finally:
        conn.close()
    return False

def get_user_role(username: Optional[str]) -> str:
    if not username:
        return "user"
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT role, is_admin FROM users WHERE LOWER(username) = LOWER(?)", (username,))
    row = cur.fetchone()
    conn.close()
    if row:
        role, is_admin = row
        if is_admin == 1 or role == "admin":
            return "admin"
        return role
    return "user"

def create_change_request(username: str, action_type: str, target_id: Optional[str] = None, payload: Optional[str] = None, description: Optional[str] = None):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO change_requests (username, action_type, target_id, payload, description, status)
        VALUES (?, ?, ?, ?, ?, 'pending')
    """, (username, action_type, target_id, payload, description))
    conn.commit()
    conn.close()

def get_candidate_name(candidate_id: int) -> str:
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT full_name FROM candidate_metadata WHERE id = ?", (candidate_id,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else f"ID {candidate_id}"

def get_job_title(job_id: int) -> str:
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT title FROM jobs WHERE id = ?", (job_id,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else f"ID {job_id}"

# ── Candidates ─────────────────────────────────────────────────────────────────
@app.get("/api/candidates")
def list_candidates(request: Request):
    username = request.headers.get("x-user-username")
    
    is_user_admin = False
    is_external = False
    role = "user"
    if username:
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        cur = conn.cursor()
        cur.execute("SELECT is_external, is_admin, role FROM users WHERE LOWER(username) = LOWER(?)", (username,))
        row = cur.fetchone()
        conn.close()
        if row:
            is_external = (row[0] == 1)
            is_user_admin = (row[1] == 1 or row[2] == "admin" or is_admin_or_hr(username))
            role = row[2]
            if is_external:
                raise HTTPException(status_code=403, detail="Forbidden")
            
    if not is_user_approved(username):
        return []

    rows = get_candidates_list(username, role="admin" if is_user_admin else "user")
    
    # Replace None values with empty string
    for row in rows:
        for k, v in row.items():
            if v is None:
                row[k] = ""

    # Mask certifications for non-admin and non-HR users
    is_user_admin_or_hr = is_admin_or_hr(username)
    if not is_user_admin_or_hr:
        for row in rows:
            row["certifications"] = "[HIDDEN]"

    if not is_admin_or_hr(username):
        keywords = get_masked_keywords()
        rows = [mask_candidate_record(row, keywords) for row in rows]
        
    rows = apply_user_hidden_fields(rows, username)
    return rows

class CustomColumn(BaseModel):
    col_key: str
    col_label: str
    description: str

@app.post("/api/columns")
def add_column(col: CustomColumn, request: Request):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    clean_key = re.sub(r'[^a-zA-Z0-9_]', '', col.col_key.replace(' ', '_')).lower()
    
    cur.execute("PRAGMA table_info(candidate_metadata)")
    existing = [c[1] for c in cur.fetchall()]
    if clean_key in existing:
        conn.close()
        raise HTTPException(status_code=400, detail="Column already exists")
        
    try:
        cur.execute(f"ALTER TABLE candidate_metadata ADD COLUMN {clean_key} TEXT")
        cur.execute("INSERT INTO custom_columns (col_key, col_label, description) VALUES (?, ?, ?)", 
                    (clean_key, col.col_label, col.description))
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
    conn.close()
    return {"status": "added", "col_key": clean_key}

@app.get("/api/columns")
def get_columns():
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT col_key, col_label FROM custom_columns")
    customs = [{"col_key": row[0], "col_label": row[1]} for row in cur.fetchall()]
    conn.close()
    
    base_cols = [
        {"col_key": "full_name", "col_label": "Name"},
        {"col_key": "candidate_status", "col_label": "Candidate Status"},
        {"col_key": "source", "col_label": "Source"},
        {"col_key": "total_experience", "col_label": "Total Exp"},
        {"col_key": "pega_experience", "col_label": "Pega Exp"},
        {"col_key": "cdh_exp", "col_label": "CDH Exp"},
        {"col_key": "ctc", "col_label": "Current CTC"},
        {"col_key": "expected_ctc", "col_label": "Exp CTC"},
        {"col_key": "percentage_hike", "col_label": "Percentage Hike"},
        {"col_key": "candidate_interview_status", "col_label": "Candidate Interview Status"},
        {"col_key": "availability_in_days", "col_label": "Availability in Days"},
        {"col_key": "notice_period", "col_label": "Notice Period"},
        {"col_key": "phone", "col_label": "Phone No"},
        {"col_key": "email", "col_label": "Email"},
        {"col_key": "linkedin", "col_label": "LinkedIn"},
        {"col_key": "current_location", "col_label": "Current Location"},
        {"col_key": "pref_locations", "col_label": "Pref Locations"},
        {"col_key": "current_organization", "col_label": "Current Employment"},
        {"col_key": "current_client", "col_label": "Current Client"},
        {"col_key": "domain", "col_label": "Domain"},
        {"col_key": "tier", "col_label": "Tier (Tier1 Tier2 Tier3)"},
        {"col_key": "certification_version", "col_label": "Certification Version"},
        {"col_key": "skills", "col_label": "Skills"},
        {"col_key": "certifications", "col_label": "Certifications"},
        {"col_key": "notescomments", "col_label": "Notes / Comments"}
    ]
    return {"base": base_cols, "custom": customs}

@app.delete("/api/columns/{col_key}")
def delete_column(col_key: str, request: Request):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM custom_columns WHERE col_key=?", (col_key,))
        try:
            cur.execute(f"ALTER TABLE candidate_metadata DROP COLUMN {col_key}")
        except Exception:
            pass # older sqlite versions might not support drop column
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
    conn.close()
    return {"status": "deleted"}

@app.put("/api/candidates/{candidate_id}")
async def update_candidate(candidate_id: int, request: Request, background_tasks: BackgroundTasks):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)

    body = await request.json()
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur  = conn.cursor()
    
    # Check permission
    cur.execute("SELECT created_by FROM candidate_metadata WHERE id = ?", (candidate_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Candidate not found")
    created_by = row[0]
    if not is_admin_or_hr(username):
        if created_by and created_by.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")

    cur.execute("PRAGMA table_info(candidate_metadata)")
    allowed_cols = [c[1] for c in cur.fetchall()]
    updates = {k: v for k, v in body.items() if k in allowed_cols and k != 'id' and v is not None and v != '[HIDDEN]'}
    
    # Server-side validation for numeric edits
    for int_col in ['notice_period', 'availability_in_days']:
        if int_col in updates and updates[int_col] != "":
            try:
                updates[int_col] = int(float(updates[int_col]))
            except ValueError:
                conn.close()
                return JSONResponse(status_code=400, content={"detail": f"{int_col} must be an integer"})
            
    for exp_col in ['total_experience', 'pega_experience', 'cdh_exp']:
        if exp_col in updates and updates[exp_col] != "":
            try:
                updates[exp_col] = float(updates[exp_col])
            except ValueError:
                conn.close()
                return JSONResponse(status_code=400, content={"detail": f"{exp_col} must be a number"})

    if not updates:
        conn.close()
        return {"status": "no changes"}



    set_clause = ", ".join(f"{k}=?" for k in updates)
    cur.execute(
        f"UPDATE candidate_metadata SET {set_clause} WHERE id=?",
        list(updates.values()) + [candidate_id]
    )
    conn.commit()
    conn.close()

    # Re-trigger matching if matching-related details have changed
    match_related_fields = {
        'full_name', 'total_experience', 'pega_experience', 'cdh_exp',
        'skills', 'certifications', 'current_location', 'pref_locations'
    }
    if any(field in updates for field in match_related_fields):
        background_tasks.add_task(match_candidate_to_all_jobs, candidate_id)

    cname = get_candidate_name(candidate_id)
    log_activity_db(username or "unknown", f"updated candidate '{cname}' details")

    return {"status": "updated"}

@app.delete("/api/candidates/{candidate_id}")
def delete_candidate(candidate_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)
    
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur  = conn.cursor()
    
    cur.execute("SELECT full_name, created_by FROM candidate_metadata WHERE id = ?", (candidate_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Candidate not found")
    cname, created_by = row
    
    if not is_admin_or_hr(username):
        if created_by and created_by.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")
            
    cur.execute("DELETE FROM job_candidates WHERE candidate_id=?", (candidate_id,))
    cur.execute("DELETE FROM candidate_metadata WHERE id=?", (candidate_id,))
    conn.commit()
    conn.close()
    log_activity_db(username or "unknown", f"deleted candidate '{cname}'")
    return {"status": "deleted"}

@app.get("/api/candidates/{candidate_id}/jobs")
def get_candidate_jobs(candidate_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # Check candidate existence and permission
    cur.execute("SELECT created_by FROM candidate_metadata WHERE id = ?", (candidate_id,))
    cand_row = cur.fetchone()
    if not cand_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Candidate not found")
        
    created_by = cand_row[0]
    if not is_admin_or_hr(username):
        if created_by and created_by.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")

    cur.execute("""
        SELECT j.*, jc.status as match_status, jc.ai_reason
        FROM jobs j
        JOIN job_candidates jc ON j.id = jc.job_id
        WHERE jc.candidate_id = ?
    """, (candidate_id,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.get("/api/candidates/{candidate_id}/formatted-resume")
def get_formatted_resume_data(candidate_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM candidate_metadata WHERE id = ?", (candidate_id,))
    row = cur.fetchone()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Candidate not found")

    candidate = dict(row)
    created_by = candidate.get("created_by")
    if not is_admin_or_hr(username):
        if created_by and created_by.lower() != username.lower():
            raise HTTPException(status_code=403, detail="Forbidden")

    # Check cache first
    if candidate.get("formatted_json"):
        try:
            return json.loads(candidate.get("formatted_json"))
        except Exception:
            pass

    filename = candidate.get("filename", "")
    
    # Try to load original text
    text = ""
    if filename:
        path = os.path.join(UPLOAD_DIR, filename)
        if os.path.exists(path) and not filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            try:
                if filename.lower().endswith(".pdf"):
                    loader = SafePyMuPDFLoader(path)
                else:
                    loader = Docx2txtLoader(path)
                docs = loader.load()
                text = "\n".join([d.page_content for d in docs])
            except Exception as e:
                print(f"Error loading resume file: {e}")

    # Fallback to metadata if file text couldn't be loaded
    if not text:
        text = f"""
        Name: {candidate.get('full_name')}
        Experience: {candidate.get('total_experience')} years (Pega: {candidate.get('pega_experience')} years, CDH: {candidate.get('cdh_exp')} years)
        Current Employment: {candidate.get('current_organization')}
        Location: {candidate.get('current_location')}
        Skills: {candidate.get('skills')}
        Certifications: {candidate.get('certifications')}
        Preferred Location: {candidate.get('pref_locations')}
        CTC: {candidate.get('ctc')}
        Expected CTC: {candidate.get('expected_ctc')}
        """

    _, llm = get_models()

    prompt = f"""You are an expert resume formatter. Extract details from the candidate resume text below and structure them into the exact JSON template provided.
Make sure to fill all fields based on the candidate's actual data. If a field is not present in the resume text, leave it blank or default appropriately.

Template Structure:
{{
  "full_name": "Name of the candidate",
  "job_title": "Determine a standard job designation based on their experience and skills, e.g. 'PEGA - CERTIFIED SENIOR SYSTEM ANALYST', 'Pega Lead Business Architect', or similar. If they are a Pega certified system architect, use a format like 'PEGA - CERTIFIED SENIOR SYSTEM ARCHITECT' or 'PEGA - CERTIFIED SYSTEM ARCHITECT'",
  "profile_summary": "A professional profile summary. You MUST format this summary using the exact wording layout: '[X]+ years of experience in [domain/industry] with a proven track record of [key achievement]. Expertise in [Skill 1], [Skill 2], [Skill 3], and [Skill 4]. Experienced in [tools/platforms] and capable of [key capability]. Adept at working with [stakeholders] to deliver [outcomes].' Fill in the bracketed placeholders using the candidate's actual data. Do not leave brackets in the output, resolve them completely.",
  "domain_skills": [
    "A concise list of 4-6 specific domain or functional skill areas, e.g. 'Pega PRPC', 'Decisioning (CDH)', 'Integration Services', 'Data Modeling', 'Agile Methodologies'"
  ],
  "technical_skills": {{
    "primary": "Primary Tool/Platform: [list primary tool(s) and version(s) if known, e.g. 'Pega PRPC: v8.x, v7.x']",
    "languages": "Languages: [list programming/database languages, e.g. 'Java, SQL, XML']",
    "frontend": "Frontend: [list frontend technologies, e.g. 'HTML5, CSS3, JavaScript, React']",
    "others": "Others: [list other tools and platforms, e.g. 'Git, Jira, Azure, Maven']"
  }},
  "education": [
    {{
      "degree": "Degree name (e.g. 'B.E.', 'MCA', 'B.Tech')",
      "field": "Field of study (e.g. 'Computer Science & Engineering')",
      "school": "University or College Name",
      "years": "Year Start - Year End or Year of Completion (e.g. '2015 - 2019')"
    }}
  ],
  "certifications": [
    "List of certifications found, e.g. 'Pega Certified Senior System Architect (CSSA)', 'Certified System Architect (CSA)', 'Certified Pega Decisioning Consultant (CPDC)'"
  ],
  "work_experience": [
    {{
      "company": "Company Name",
      "dates": "Start Date - End Date (e.g. 'Jul 2019 - Present')",
      "role": "Role or Job Title",
      "bullets": [
        "Accomplishment or key responsibility bullet point (limit to 3-5 high-impact bullets per job)"
      ]
    }}
  ],
  "recognitions": [
    {{
      "date": "Month Year / Year (e.g. '2022' or 'Dec 2021')",
      "description": "Award title or recognition detail (e.g. 'Star of the Quarter for excellent project delivery')"
    }}
  ]
}}

Rules:
1. Return ONLY the valid JSON block. No explanation, no markdown backticks, no markdown blocks. Just raw valid JSON.
2. If certifications are empty, extract them from the text (do not use "[HIDDEN]").
3. All fields must be clean strings, numbers, arrays, or objects as defined in the template.

Resume Text:
{text[:8000]}

JSON:"""

    try:
        resp = llm.invoke([HumanMessage(content=prompt)])
        raw = resp.content.strip()
        if "```json" in raw:
            raw = raw.split("```json")[1].split("```")[0].strip()
        elif "```" in raw:
            raw = raw.split("```")[1].split("```")[0].strip()
        start, end = raw.find('{'), raw.rfind('}')
        if start != -1 and end != -1:
            raw = raw[start:end+1]
        data = json.loads(raw)

        # Cache the result in DB
        try:
            conn = sqlite3.connect(STATS_DB, timeout=30.0)
            cur = conn.cursor()
            cur.execute("UPDATE candidate_metadata SET formatted_json = ? WHERE id = ?", (json.dumps(data), candidate_id))
            conn.commit()
            conn.close()
        except Exception as e_cache:
            print(f"Error caching formatted resume: {e_cache}")

    except Exception as e:
        print(f"Error parsing resume via LLM: {e}")
        # Return fallback structured data
        data = {
            "full_name": candidate.get("full_name"),
            "job_title": "Pega Professional",
            "profile_summary": f"{candidate.get('total_experience', 0)} years of experience in IT with skills in {candidate.get('skills', '')}.",
            "domain_skills": [s.strip() for s in str(candidate.get("skills", "")).split(",") if s.strip()][:4],
            "technical_skills": {
                "primary": "Primary Tool/Platform: Pega",
                "languages": "Languages: Java, SQL",
                "frontend": "Frontend: HTML, CSS, JavaScript",
                "others": "Others: Git, Jira"
            },
            "education": [],
            "certifications": [c.strip() for c in str(candidate.get("certifications", "")).split(",") if c.strip()],
            "work_experience": [
                {
                    "company": candidate.get("current_organization") or "Current Employer",
                    "dates": "N/A",
                    "role": "Pega Developer",
                    "bullets": ["Contributed to application development and configuration."]
                }
            ],
            "recognitions": []
        }

    return data

@app.put("/api/candidates/{candidate_id}/formatted-resume")
async def update_formatted_resume_data(candidate_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    
    body = await request.json()
    
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    try:
        cur = conn.cursor()
        cur.execute("SELECT created_by FROM candidate_metadata WHERE id = ?", (candidate_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Candidate not found")
            
        created_by = row[0]
        if not is_admin_or_hr(username):
            if created_by and created_by.lower() != username.lower():
                raise HTTPException(status_code=403, detail="Forbidden")
                
        cur.execute("UPDATE candidate_metadata SET formatted_json = ? WHERE id = ?", (json.dumps(body), candidate_id))
        conn.commit()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
    
    cname = get_candidate_name(candidate_id)
    log_activity_db(username or "unknown", f"updated formatted resume for candidate '{cname}'")
    return {"status": "updated"}

# ── Upload & Extract ────────────────────────────────────────────────────────────
EXTRACT_PROMPT = """You are an expert resume and email parser. Extract EVERY piece of information from the resume and the email message (if provided) below.
Be extremely thorough — search all sections: Summary, Experience, Skills, Education, Certifications, Contact, Headers, Footers, and the email body.

Map the extracted data according to the database keys and their corresponding column headings/labels.

Return ONLY a valid JSON object with these exact keys (no extra text, no markdown, just raw JSON):

{{
  "full_name": "<Full name from top of resume. Matching column heading: 'Name'>",
  "total_experience": <number — total years of professional work experience. MUST be extracted from the resume document itself (do not take this from the email body). Calculate if needed. 0 if fresher. Matching column heading: 'Total Exp'>,
  "pega_experience": <number — years of Pega PRPC/BPM experience specifically. If the email body specifies Pega experience (e.g. '5 yrs into pega'), MUST extract this value from the email body. 0 if none. Matching column heading: 'Pega Exp'>,
  "cdh_exp": <number — years of Pega Customer Decision Hub (CDH) experience specifically. If the email body specifies CDH experience, take that. Otherwise, calculate it from the resume (look at durations of CDH-related projects or roles). 0 if none. Matching column heading: 'CDH Exp'>,
  "ctc": "<Current CTC / current salary / current package. E.g. '8 LPA', '10.5 Lacs', '900000'. Look for salary, CTC, package, LPA, Lakhs. Empty if not found. Matching column heading: 'Current CTC'>",
  "expected_ctc": "<Expected CTC / expected salary / expected package. E.g. '12 LPA', '15 Lacs'. Look for expected CTC, ECTC, expected salary, expected package, expectation. Empty if not found. Matching column heading: 'Exp CTC'>",
  "percentage_hike": "<Expected percentage hike. Calculate if both current and expected CTC are present, else empty. Matching column heading: 'Percentage Hike'>",
  "candidate_interview_status": "<Default to empty string '' unless resume has recruiter notes on status. Matching column heading: 'Candidate Interview Status'>",
  "availability_in_days": <number — integer representing days until available. e.g. for 'immediate' use 0, '1 month' use 30. Return null if not found. Matching column heading: 'Availability in Days'>,
  "notice_period": "<Notice period in integer DAYS ONLY. E.g. for 'Immediate' return 0. For '1 month' return 30. ONLY output digits, no text. Matching column heading: 'Notice Period'>",
  "phone": "<Phone number with country code. Include only digits and + sign. Matching column heading: 'Phone No'>",
  "email": "<Email address. Matching column heading: 'Email'>",
  "linkedin": "<Full LinkedIn profile URL if found, else empty string. Matching column heading: 'LinkedIn'>",
  "current_location": "<Current city/location. Matching column heading: 'Current Location'>",
  "pref_locations": "<Preferred locations to work. Matching column heading: 'Pref Locations'>",
  "current_organization": "<Current or most recent employer name. Matching column heading: 'Current Employment'>",
  "current_client": "<Current client working for, if mentioned. Otherwise empty. Matching column heading: 'Current Client'>",
  "domain": "<Domain of expertise, e.g. Banking, Telecom, Healthcare. Extract from recent projects. Matching column heading: 'Domain'>",
  "tier": "<Tier1, Tier2, or Tier3 based on their college/university or companies. Guess if possible, or empty. Matching column heading: 'Tier (Tier1 Tier2 Tier3)'>",
  "certification_version": "<Version of Pega certifications, e.g. 8.6, 8.8. Empty if not found. Matching column heading: 'Certification Version'>",
  "skills": "<All technical skills comma-separated. Matching column heading: 'Skills'>",
  "certifications": "<All certifications and courses comma-separated. Matching column heading: 'Certifications'>",
  "source": "<Extract source of candidate profile if mentioned in resume, e.g. 'LinkedIn', 'Naukri', 'Resume Upload'. If none is found, return 'Resume Upload'. Matching column heading: 'Source'>"{custom_fields}
}}

Rules:
- cdh_exp, pega_experience, total_experience: Must be purely numerical floats or 0.
- notice_period: Must be strictly an integer (0, 15, 30, 60). No strings.
- availability_in_days: Integer number of days.
- If a field is not found, use an empty string "" (or null for numbers). Do not invent data.
- If both a resume and an email message are provided, extract and merge the details. The email message may contain more recent or specific details (such as current/expected CTC, notice period, or Pega/CDH experience); prioritize information from the email message in case of conflict. However, the total_experience (total years of experience) MUST always be extracted from the resume document itself (not the email body), as the resume lists the full work history.

Resume and Email Content:
{text}

JSON:"""


def process_resume_logic(safe_name: str, path: str, is_approved: int = 1, username: str = "unknown", email_message: str = None, sender_email: str = None):
    try:
        _, llm = get_models()
        embeddings, _ = get_models()

        # Load document
        if safe_name.lower().endswith(".pdf"):
            loader = SafePyMuPDFLoader(path)
        else:
            loader = Docx2txtLoader(path)
        docs = loader.load()
        text = "\n".join([d.page_content for d in docs])

        # Fetch custom columns for the prompt
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        cur = conn.cursor()
        cur.execute("SELECT col_key, col_label, description FROM custom_columns")
        custom_cols = cur.fetchall()
        conn.close()
        
        custom_fields_str = ""
        if custom_cols:
            for col_key, col_label, desc in custom_cols:
                # Include the exact column heading / label and description to make extraction precise
                desc_str = f"Extract data corresponding to column heading '{col_label}'"
                if desc:
                    desc_str += f" ({desc})"
                custom_fields_str += f',\n  "{col_key}": "<{desc_str}>"'
                
        combined_text = text[:7000]
        if email_message:
            combined_text += f"\n\n=== EMAIL MESSAGE BODY ===\n{email_message}\n=========================="
            
        prompt_str = EXTRACT_PROMPT.format(text=combined_text, custom_fields=custom_fields_str)
        
        # Add a simple retry mechanism for rate limits
        max_retries = 3
        resp = None
        for attempt in range(max_retries):
            try:
                resp = llm.invoke([HumanMessage(content=prompt_str)])
                break
            except Exception as api_err:
                if "429" in str(api_err) and attempt < max_retries - 1:
                    import time
                    time.sleep(3) # Wait 3 seconds before retrying
                    continue
                raise api_err
                
        if resp is None:
            raise Exception("Failed to get response from AI model")

        raw  = resp.content.strip()

        if "```json" in raw:
            raw = raw.split("```json")[1].split("```")[0].strip()
        elif "```" in raw:
            raw = raw.split("```")[1].split("```")[0].strip()
        start, end = raw.find('{'), raw.rfind('}')
        if start != -1 and end != -1:
            raw = raw[start:end+1]

        data = json.loads(raw)
        if username == "email_worker":
            data['source'] = 'Import from Mail'
        if email_message:
            data['email_message'] = email_message
        if sender_email:
            data['sender_email'] = sender_email

        # -- Start Data Validation & Normalization --
        import re
        
        # Phone: Keep only digits and +
        if 'phone' in data and data['phone']:
            data['phone'] = re.sub(r'[^\d+]', '', str(data['phone']))
            
        # Email: Basic validation
        if 'email' in data and data['email']:
            if '@' not in str(data['email']):
                data['email'] = ""
                
        # Experience: Force float
        for exp_field in ['total_experience', 'pega_experience', 'cdh_exp']:
            if exp_field in data and data[exp_field] not in [None, ""]:
                try:
                    match = re.search(r'\d+(\.\d+)?', str(data[exp_field]))
                    data[exp_field] = float(match.group()) if match else 0.0
                except Exception:
                    data[exp_field] = 0.0
                    
        # Integer fields
        for num_field in ['notice_period', 'availability_in_days']:
            if num_field in data and data[num_field] not in [None, ""]:
                np_str = str(data[num_field]).lower()
                if 'immediate' in np_str:
                    data[num_field] = 0
                else:
                    try:
                        match = re.search(r'\d+', np_str)
                        val = int(match.group()) if match else ""
                        if match and 'month' in np_str:
                            val = val * 30
                        data[num_field] = val
                    except Exception:
                        data[num_field] = ""
        # -- End Data Validation & Normalization --

        # Check for existing match (excluding the placeholder we just created)
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        cur = conn.cursor()
        
        # Match by sender_email first if provided
        match_id = None
        if sender_email:
            cur.execute("SELECT id FROM candidate_metadata WHERE LOWER(sender_email) = ? OR LOWER(email) = ? ORDER BY id DESC LIMIT 1", (sender_email.lower(), sender_email.lower()))
            row = cur.fetchone()
            if row:
                match_id = row[0]
                
        if not match_id:
            cur.execute("SELECT id, full_name, email, phone, sender_email FROM candidate_metadata WHERE (filename != ? OR filename IS NULL)", (safe_name,))
            existing_candidates = [
                {
                    "id": r[0],
                    "full_name": r[1] or "",
                    "email": r[2] or "",
                    "phone": r[3] or "",
                    "sender_email": r[4] or ""
                }
                for r in cur.fetchall()
            ]
            
            email = data.get('email', '')
            phone = data.get('phone', '')
            name = data.get('full_name', '')
            
            norm_email = normalize_email(email)
            norm_sender_email = normalize_email(sender_email)
            
            for ec in existing_candidates:
                ec_email = normalize_email(ec["email"])
                ec_sender = normalize_email(ec["sender_email"])
                
                if norm_email and (norm_email == ec_email or norm_email == ec_sender):
                    match_id = ec["id"]
                    break
                if norm_sender_email and (norm_sender_email == ec_email or norm_sender_email == ec_sender):
                    match_id = ec["id"]
                    break
                if phone and ec["phone"] and phones_match(phone, ec["phone"]):
                    match_id = ec["id"]
                    break
                if name and ec["full_name"] and is_similar_name(name, ec["full_name"]):
                    match_id = ec["id"]
                    break
                
        if match_id:
            # Match found! Delete the placeholder record we created in upload_resume
            cur.execute("DELETE FROM candidate_metadata WHERE filename = ?", (safe_name,))
            
            # Fetch the existing candidate metadata to merge values
            cur.execute("PRAGMA table_info(candidate_metadata)")
            allowed_cols = {c[1] for c in cur.fetchall()}
            
            cur.execute("SELECT * FROM candidate_metadata WHERE id = ?", (match_id,))
            cur.row_factory = sqlite3.Row
            existing_row = dict(cur.fetchone())
            
            updates = {}
            for k, v in data.items():
                if k in allowed_cols and k != 'id' and k != 'filename':
                    if v is not None and v != "":
                        existing_val = existing_row.get(k)
                        if username == "email_worker" or existing_val is None or str(existing_val).strip() == "" or existing_val == 0.0 or existing_val == 0:
                            updates[k] = v
            # Always update or attach the filename to the matched candidate
            updates['filename'] = safe_name
            if username == "email_worker":
                updates['source'] = 'Import from Mail'
                if email_message:
                    updates['email_message'] = email_message
                if sender_email:
                    updates['sender_email'] = sender_email
            
            if updates:
                set_clause = ", ".join(f"{k}=?" for k in updates)
                cur.execute(f"UPDATE candidate_metadata SET {set_clause} WHERE id=?", list(updates.values()) + [match_id])
            
            candidate_id = match_id
            print(f"INFO: Auto-attached uploaded resume {safe_name} to existing candidate profile ID {match_id} ({data.get('full_name', '')})")
            conn.commit()
            conn.close()
        else:
            conn.close()
            data['filename'] = safe_name
            data['candidate_status'] = 'New'
            data['is_approved'] = is_approved
            data['created_by'] = username
            candidate_id = log_candidate(data)
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        error_msg = str(e)[:100]
        print(f"Error processing resume {safe_name}: {error_msg}")
        try:
            conn = sqlite3.connect(STATS_DB, timeout=30.0)
            cur = conn.cursor()
            cur.execute("DELETE FROM candidate_metadata WHERE filename = ?", (safe_name,))
            conn.commit()
            conn.close()
        except Exception as db_err:
            print(f"Error deleting placeholder from DB: {db_err}")
            
        if os.path.exists(path):
            try:
                os.remove(path)
                print(f"Cleaned up failed resume file: {path}")
            except Exception as file_err:
                print(f"Error cleaning up failed resume file {path}: {file_err}")
        return

    # Add to ChromaDB
    try:
        for d in docs:
            d.metadata['source'] = safe_name
        splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
        chunks   = splitter.split_documents(docs)
        Chroma.from_documents(chunks, embeddings, persist_directory=CHROMA_PATH)
    except Exception as e:
        pass

    # Automatically match this candidate to all active JDs in the database
    if candidate_id and is_approved == 1:
        match_candidate_to_all_jobs(candidate_id)

def match_candidate_to_all_jobs(candidate_id: int):
    try:
        # Query candidate details
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM candidate_metadata WHERE id = ?", (candidate_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return
        data = dict(row)
        
        # Query all jobs (JDs)
        cand_creator = data.get('created_by')
        if cand_creator and cand_creator.lower() != "admin":
            cur.execute("SELECT id, title, description FROM jobs WHERE LOWER(created_by) = LOWER(?)", (cand_creator,))
        else:
            cur.execute("SELECT id, title, description FROM jobs")
        jobs = [dict(r) for r in cur.fetchall()]
        conn.close()

        if not jobs:
            return

        # Format JDs for the LLM
        jds_str_list = []
        for job in jobs:
            jds_str_list.append(f"JD ID: {job['id']}\nTitle: {job['title']}\nDescription: {job['description']}\n---")
        formatted_jds = "\n".join(jds_str_list)

        _, llm = get_models()

        prompt = f"""You are an expert technical recruiter. Evaluate candidate "{data.get('full_name', 'Candidate')}" against the following Job Descriptions "pin to pin".

Candidate Details:
- Skills: {data.get('skills', '')}
- Experience: {data.get('total_experience', 0)} years (Pega Experience: {data.get('pega_experience', 0)} years, CDH Experience: {data.get('cdh_exp', 0)} years)
- Certifications: {data.get('certifications', '')}
- Current Location: {data.get('current_location', '')}
- Preferred Locations: {data.get('pref_locations', '')}

Job Descriptions:
{formatted_jds}

Rules for evaluation:
1. Numeric Experience Matching: If a Job Description asks for "X+ years of experience", a candidate matches if their experience is greater than or equal to X.
   - For example: if Job Description requires "1+ years of experience in pega", then candidates with 3.0 years, 4.0 years, or 4.8 years of Pega experience all match perfectly because 3.0 >= 1.0, 4.0 >= 1.0, and 4.8 >= 1.0.
2. Certification Abbreviations:
   - CSSA is equivalent to any of: "PEGA Certified Senior System Architect", "Pega Certified Senior System Architect", "Certified Pega Senior System Architect", "Senior System Architect", or "CSSA".
   - CSA is equivalent to any of: "PEGA Certified System Architect", "Pega Certified System Architect", "Certified Pega System Architect", "System Architect", or "CSA".
   - LSA is equivalent to any of: "PEGA Certified Lead System Architect", "Pega Certified Lead System Architect", "Certified Pega Lead System Architect", "Lead System Architect", or "LSA".
3. Do not invent requirements. If the Job Description only mentions Pega experience, do NOT reject candidates for lacking CSSA or other unrelated certifications.
4. Location Matching: If the Job Description specifies a location requirement, a candidate matches if their Current Location or any of their Preferred Locations match the specified job location (e.g. if the Job Description mentions 'Chennai', a candidate with Current Location or Preferred Location 'Chennai' is a match).

For each Job Description, decide if the candidate is a good match based on the rules.
Respond with a JSON list of objects for matches only:
[
  {{"job_id": <job_id>, "reason": "<1-sentence explanation of fit based on specific experience, skills, and location>"}}
]
If there are no matches, return an empty list [].
Return ONLY the JSON block, no markdown, no other text."""

        # Add a retry mechanism for rate limits
        max_retries = 3
        resp = None
        for attempt in range(max_retries):
            try:
                resp = llm.invoke([HumanMessage(content=prompt)])
                break
            except Exception as api_err:
                if "429" in str(api_err) and attempt < max_retries - 1:
                    import time
                    time.sleep(3) # Wait 3 seconds before retrying
                    continue
                raise api_err
                
        if resp is None:
            raise Exception("Failed to get response from AI model")

        raw  = resp.content.strip()

        if "```json" in raw:
            raw = raw.split("```json")[1].split("```")[0].strip()
        elif "```" in raw:
            raw = raw.split("```")[1].split("```")[0].strip()
        start, end = raw.find('['), raw.rfind(']')
        if start != -1 and end != -1:
            raw = raw[start:end+1]

        matches = json.loads(raw)
        
        conn = sqlite3.connect(STATS_DB, timeout=30.0)
        cur = conn.cursor()
        
        # Clear existing automatic matches for this candidate
        cur.execute("DELETE FROM job_candidates WHERE candidate_id = ? AND status = 'matched'", (candidate_id,))
        
        matched_any = False
        if matches:
            for match in matches:
                job_id = match.get("job_id")
                reason = match.get("reason", "Matched based on candidate details.")
                if job_id:
                    cur.execute("""
                        INSERT INTO job_candidates (job_id, candidate_id, ai_reason, status) 
                        VALUES (?, ?, ?, 'matched')
                        ON CONFLICT(job_id, candidate_id) DO UPDATE SET ai_reason = excluded.ai_reason
                    """, (job_id, candidate_id, reason))
                    matched_any = True
        
        # Automatically update qualification status
        cur.execute("SELECT COUNT(*) FROM job_candidates WHERE candidate_id = ?", (candidate_id,))
        cnt = cur.fetchone()[0]
        is_qualified = 1 if cnt > 0 else 0
        cur.execute("UPDATE candidate_metadata SET is_qualified = ? WHERE id = ?", (is_qualified, candidate_id))
        
        conn.commit()
        conn.close()
    except Exception as match_err:
        print(f"Error auto-matching candidate {candidate_id} to jobs: {match_err}")

def process_excel_file_logic(safe_name: str, path: str, username: str):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        
        # Fetch custom columns
        cur.execute("SELECT col_key, col_label FROM custom_columns")
        custom_cols = {row[1].strip().lower(): row[0] for row in cur.fetchall()}
        
        # Mappings for common columns
        column_mappings = {
            'name': 'full_name',
            'full name': 'full_name',
            'candidate name': 'full_name',
            'candidate': 'full_name',
            'names': 'full_name',
            
            'phone': 'phone',
            'phone no': 'phone',
            'phone number': 'phone',
            'mobile': 'phone',
            'mobile number': 'phone',
            'mobile no': 'phone',
            'contact': 'phone',
            'contact number': 'phone',
            'contact no': 'phone',
            
            'email': 'email',
            'email id': 'email',
            'email address': 'email',
            
            'skills': 'skills',
            'key skills': 'skills',
            'technical skills': 'skills',
            
            'experience': 'total_experience',
            'exp': 'total_experience',
            'total exp': 'total_experience',
            'total experience': 'total_experience',
            'work experience': 'total_experience',
            
            'pega experience': 'pega_experience',
            'pega exp': 'pega_experience',
            
            'cdh experience': 'cdh_exp',
            'cdh exp': 'cdh_exp',
            
            'current ctc': 'ctc',
            'ctc': 'ctc',
            'salary': 'ctc',
            
            'expected ctc': 'expected_ctc',
            
            'percentage hike': 'percentage_hike',
            'hike': 'percentage_hike',
            
            'notice period': 'notice_period',
            'np': 'notice_period',
            
            'current location': 'current_location',
            'location': 'current_location',
            
            'preferred locations': 'pref_locations',
            'preferred location': 'pref_locations',
            'pref locations': 'pref_locations',
            
            'current organization': 'current_organization',
            'current employer': 'current_organization',
            'employer': 'current_organization',
            'current employment': 'current_organization',
            
            'current client': 'current_client',
            
            'domain': 'domain',
            'tier': 'tier',
            'certification version': 'certification_version',
            'certifications': 'certifications',
            
            # Additional maps for thorough Excel parsing
            'linkedin': 'linkedin',
            'linkedin url': 'linkedin',
            'linkedin profile': 'linkedin',
            'notes': 'notescomments',
            'comments': 'notescomments',
            'notes/comments': 'notescomments',
            'notescomments': 'notescomments',
            'feedback': 'notescomments',
            'candidate status': 'candidate_status',
            'candidate_status': 'candidate_status',
            'status': 'candidate_status',
            'candidate interview status': 'candidate_interview_status',
            'interview status': 'candidate_interview_status',
            'availability': 'availability_in_days',
            'availability in days': 'availability_in_days',
            'availability_in_days': 'availability_in_days',
            'source': 'source',
            'candidate source': 'source',
            'how did you find us': 'source'
        }
        
        # Load existing candidates from DB
        cur.execute("SELECT id, full_name, email, phone FROM candidate_metadata WHERE LOWER(created_by) = LOWER(?)", (username,))
        existing_candidates = [
            {
                "id": r[0],
                "full_name": r[1] or "",
                "email": r[2] or "",
                "phone": r[3] or ""
            }
            for r in cur.fetchall()
        ]
        
        # Iterate over all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Find the header row (iterating up to 10 rows)
            header_row_idx = None
            headers = []
            for r_idx in range(1, min(ws.max_row + 1, 11)):
                row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, ws.max_column + 1)]
                non_empty = [v for v in row_vals if v is not None]
                if len(non_empty) >= 2:
                    is_header = False
                    for val in non_empty:
                        val_str = str(val).strip().lower()
                        if val_str in column_mappings or val_str in custom_cols:
                            is_header = True
                            break
                    if is_header:
                        header_row_idx = r_idx
                        headers = [str(h).strip() if h is not None else "" for h in row_vals]
                        break
            
            if not header_row_idx:
                header_row_idx = 1
                headers = [str(ws.cell(row=1, column=c_idx).value).strip() if ws.cell(row=1, column=c_idx).value is not None else "" for c_idx in range(1, ws.max_column + 1)]
            
            # Map headers to DB columns
            mapped_cols = {}
            for idx, h in enumerate(headers):
                h_lower = h.lower()
                if h_lower in custom_cols:
                    mapped_cols[idx] = custom_cols[h_lower]
                elif h_lower in column_mappings:
                    mapped_cols[idx] = column_mappings[h_lower]
            
            # Process rows starting from header_row_idx + 1
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                row_data = {}
                is_row_empty = True
                for col_idx in range(1, len(headers) + 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        is_row_empty = False
                    if (col_idx - 1) in mapped_cols:
                        db_col = mapped_cols[col_idx - 1]
                        row_data[db_col] = str(val).strip() if val is not None else ""
                
                if is_row_empty:
                    continue
                
                # Check that we have at least one identifying field
                if not row_data.get('full_name') and not row_data.get('email') and not row_data.get('phone'):
                    continue
                
                name = row_data.get('full_name', '')
                email = row_data.get('email', '')
                phone = row_data.get('phone', '')
                
                norm_email = normalize_email(email)
                norm_phone = normalize_phone(phone)
                
                match_id = None
                for ec in existing_candidates:
                    ec_email = normalize_email(ec["email"])
                    
                    if norm_email and norm_email == ec_email:
                        match_id = ec["id"]
                        break
                    if phone and ec["phone"] and phones_match(phone, ec["phone"]):
                        match_id = ec["id"]
                        break
                    if name and ec["full_name"] and is_similar_name(name, ec["full_name"]):
                        match_id = ec["id"]
                        break
                
                # Normalize numeric/experience fields
                if 'total_experience' in row_data and row_data['total_experience'] != "":
                    try:
                        row_data['total_experience'] = float(row_data['total_experience'])
                    except ValueError:
                        row_data['total_experience'] = 0.0
                if 'pega_experience' in row_data and row_data['pega_experience'] != "":
                    try:
                        row_data['pega_experience'] = float(row_data['pega_experience'])
                    except ValueError:
                        row_data['pega_experience'] = 0.0
                if 'cdh_exp' in row_data and row_data['cdh_exp'] != "":
                    try:
                        row_data['cdh_exp'] = float(row_data['cdh_exp'])
                    except ValueError:
                        row_data['cdh_exp'] = 0.0
                if 'notice_period' in row_data and row_data['notice_period'] != "":
                    try:
                        digits = "".join(c for c in row_data['notice_period'] if c.isdigit())
                        row_data['notice_period'] = int(digits) if digits else ""
                    except Exception:
                        row_data['notice_period'] = ""
                if 'availability_in_days' in row_data and row_data['availability_in_days'] != "":
                    try:
                        digits = "".join(c for c in row_data['availability_in_days'] if c.isdigit())
                        row_data['availability_in_days'] = int(digits) if digits else ""
                    except Exception:
                        row_data['availability_in_days'] = ""
                
                cur.execute("PRAGMA table_info(candidate_metadata)")
                allowed_cols = {c[1] for c in cur.fetchall()}
                db_data = {k: v for k, v in row_data.items() if k in allowed_cols and k != 'id'}
                if 'source' not in db_data or not db_data['source']:
                    db_data['source'] = 'Excel Import'
                
                if match_id:
                    # Update existing record
                    set_clause = ", ".join(f"{k}=?" for k in db_data)
                    cur.execute(f"UPDATE candidate_metadata SET {set_clause} WHERE id=?", list(db_data.values()) + [match_id])
                    candidate_id = match_id
                else:
                    # Insert new record
                    db_data['is_approved'] = 1
                    db_data['candidate_status'] = 'New'
                    db_data['filename'] = ""
                    db_data['created_by'] = username
                    cols_list = list(db_data.keys())
                    vals_list = list(db_data.values())
                    cur.execute(
                        f"INSERT INTO candidate_metadata ({','.join(cols_list)}) VALUES ({','.join(['?']*len(cols_list))})",
                        vals_list
                    )
                    candidate_id = cur.lastrowid
                    existing_candidates.append({
                        "id": candidate_id,
                        "full_name": name,
                        "email": email,
                        "phone": phone
                    })
                
                # Match candidate to all jobs
                match_candidate_to_all_jobs(candidate_id)
                
    except Exception as e:
        print(f"Error parsing Excel file {safe_name}: {e}")
        if os.path.exists(path):
            try:
                os.remove(path)
                print(f"Cleaned up failed Excel file: {path}")
            except Exception as file_err:
                print(f"Error cleaning up failed Excel file {path}: {file_err}")
    finally:
        try:
            cur.execute("DELETE FROM candidate_metadata WHERE filename = ? AND full_name LIKE '⏳ Parsing Excel:%'", (safe_name,))
            conn.commit()
        except Exception as db_err:
            print(f"Error clearing Excel placeholder or committing in finally: {db_err}")
        try:
            conn.close()
        except Exception:
            pass

def process_excel_file(safe_name: str, path: str, username: str = "unknown"):
    with _processing_lock:
        process_excel_file_logic(safe_name, path, username)

def process_resume(safe_name: str, path: str, is_approved: int = 1, username: str = "unknown", email_message: str = None, sender_email: str = None):
    # Use a lock to ensure only one resume is processed at a time
    # This prevents Render memory crashes (OOM), Groq Rate Limits, and SQLite database locks
    with _processing_lock:
        process_resume_logic(safe_name, path, is_approved, username, email_message, sender_email)


@app.post("/api/upload")
async def upload_resume(request: Request, background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    is_approved = 1

    # Save file
    safe_name = file.filename
    path = os.path.join(UPLOAD_DIR, safe_name)
    with open(path, "wb") as f:
        content = await file.read()
        f.write(content)

    ext = os.path.splitext(safe_name.lower())[1]
    if ext in ['.xlsx', '.xls', '.csv']:
        log_activity_db(username or "unknown", f"uploaded candidate excel sheet '{safe_name}'")
        background_tasks.add_task(process_excel_file, safe_name, path, username or "unknown")
        return {"status": "processing", "message": "Excel sheet uploaded and is processing in the background."}
    else:
        # Placeholder while processing in background
        log_candidate({"filename": safe_name, "full_name": f"⏳ Processing: {safe_name}", "is_approved": is_approved, "created_by": username or "unknown"})
        
        # Process asynchronously
        log_activity_db(username or "unknown", f"uploaded resume '{safe_name}'")
        background_tasks.add_task(process_resume, safe_name, path, is_approved, username or "unknown")

        return {"status": "processing", "message": "Resume uploaded and is processing in the background."}

# ── Chat ──────────────────────────────────────────────────────────────────────
class ChatRequest(BaseModel):
    message: str

CONVERSATIONAL_KW = [
    "hi", "hello", "hey", "how are you", "what's up",
    "good morning", "good afternoon", "good evening",
    "thanks", "thank you", "bye", "goodbye", "who are you",
]

def _is_conversational(msg: str) -> bool:
    m = msg.strip().lower()
    if m in CONVERSATIONAL_KW:
        return True
    # Short messages (≤4 words) with no data keywords → conversational
    data_kw = ["candidate", "resume", "experience", "pega", "skill", "cert",
               "ctc", "notice", "company", "email", "phone", "show", "list",
               "find", "give", "who", "which", "how many", "count", "year",
               "join", "immediate", "work", "hire", "select"]
    if len(m.split()) <= 4 and not any(k in m for k in data_kw):
        return True
    return False

def extract_search_filters(prompt: str, llm) -> dict:
    filter_extraction_prompt = f"""You are an HR database query assistant.
Extract search parameters from the following HR question into a JSON object.

HR Question: "{prompt}"

JSON Schema:
{{
  "name_query": "string or null (extract a candidate name if the user is asking about a specific person, e.g., 'Naresh' or 'Gopinath')",
  "skills_keywords": "array of strings or null (extract skills or certifications, e.g., ['pega', 'cssa', 'csa', 'lsa', 'clda'])",
  "min_pega_exp": "number or null (minimum years of Pega experience requested)",
  "min_total_exp": "number or null (minimum years of total experience requested)",
  "notice_period_max": "number or null (maximum notice period in days)",
  "current_location": "string or null (location name, e.g., 'Hyderabad')"
}}

Respond ONLY with the JSON object. Do not include any explanation or markdown formatting."""

    try:
        resp = llm.invoke([HumanMessage(content=filter_extraction_prompt)])
        ans = resp.content.strip()
        if ans.startswith("```json"):
            ans = ans[7:]
        if ans.endswith("```"):
            ans = ans[:-3]
        return json.loads(ans.strip())
    except Exception as e:
        print(f"Error extracting search filters: {e}")
        return {}

def query_candidates_by_filters(filters: dict, username: str, role: str) -> list:
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    query = "SELECT * FROM candidate_metadata WHERE 1=1"
    params = []
    
    # Non-admins and non-HRs can only see their own candidates
    if not is_admin_or_hr(username) and username:
        query += " AND LOWER(created_by) = LOWER(?)"
        params.append(username)
        
    name_q = filters.get("name_query")
    if name_q:
        query += " AND (full_name LIKE ? OR current_organization LIKE ? OR email LIKE ?)"
        params.append(f"%{name_q}%")
        params.append(f"%{name_q}%")
        params.append(f"%{name_q}%")
        
    skills = filters.get("skills_keywords")
    if skills and isinstance(skills, list):
        for sk in skills:
            if sk.strip():
                query += " AND (skills LIKE ? OR certifications LIKE ?)"
                params.append(f"%{sk.strip()}%")
                params.append(f"%{sk.strip()}%")
            
    min_pega = filters.get("min_pega_exp")
    if min_pega is not None:
        try:
            query += " AND pega_experience >= ?"
            params.append(float(min_pega))
        except ValueError:
            pass
            
    min_tot = filters.get("min_total_exp")
    if min_tot is not None:
        try:
            query += " AND total_experience >= ?"
            params.append(float(min_tot))
        except ValueError:
            pass

    max_notice = filters.get("notice_period_max")
    if max_notice is not None:
        try:
            query += " AND (CAST(notice_period AS INTEGER) <= ? OR notice_period LIKE '%immediate%')"
            params.append(int(max_notice))
        except ValueError:
            pass

    loc = filters.get("current_location")
    if loc:
        query += " AND (current_location LIKE ? OR pref_locations LIKE ?)"
        params.append(f"%{loc}%")
        params.append(f"%{loc}%")

    query += " ORDER BY timestamp DESC"
    
    try:
        cur.execute(query, params)
        rows = [dict(r) for r in cur.fetchall()]
    except Exception as e:
        print(f"Error running database query: {e}")
        rows = []
    finally:
        conn.close()
    return rows


def _find_matching_rows(rows: list, names: list) -> list:
    """Return matching exact DB rows for the given candidate names."""
    lower_names = [n.strip().lower() for n in names]
    
    # Try exact match first
    matched = [r for r in rows if r.get('full_name', '').strip().lower() in lower_names]
    
    if not matched:
        # Try partial match
        matched = []
        for r in rows:
            fn_lower = r.get('full_name', '').strip().lower()
            if any((ln in fn_lower or fn_lower in ln) and ln != '' for ln in lower_names):
                matched.append(r)
                
    records = []
    for r in matched:
        records.append({
            "name":             r.get("full_name", ""),
            "total_experience": r.get("total_experience", 0),
            "pega_experience":  r.get("pega_experience", 0),
            "skills":           r.get("skills", ""),
            "certifications":   r.get("certifications", ""),
            "ctc":              r.get("ctc", ""),
            "notice_period":    r.get("notice_period", ""),
            "organization":     r.get("current_organization", ""),
            "email":            r.get("email", ""),
            "phone":            r.get("phone", ""),
        })
    return records

@app.post("/api/chat")
def chat(body: ChatRequest, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        return {
            "type": "text",
            "answer": "Your account access is currently pending administrator approval. Please contact your system administrator to view and query candidate data."
        }

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT is_admin, role FROM users WHERE LOWER(username) = LOWER(?)", (username,))
    user_row = cur.fetchone()
    conn.close()
    is_user_admin = False
    if user_row:
        is_user_admin = (user_row[0] == 1 or user_row[1] == "admin" or is_admin_or_hr(username))

    global _embeddings, _llm, _models_loading
    if _embeddings is None or _llm is None:
        if _models_loading:
            return {"type": "text", "answer": "⏳ Hire AI is currently warming up and downloading models. This may take a few minutes depending on your internet connection. Please try again shortly!"}
        else:
            threading.Thread(target=get_models, daemon=True).start()
            return {"type": "text", "answer": "⏳ Hire AI is starting its engines... Please try again in a few seconds."}

    embeddings, llm = _embeddings, _llm
    prompt  = body.message.strip()
    p_lower = prompt.lower()

    # ── Route 1: Conversational ──────────────────────────────────────────────
    if _is_conversational(p_lower):
        resp = llm.invoke([HumanMessage(content=(
            "You are Hire AI, an intelligent HR recruitment assistant for Alamaticz Solutions. "
            "Respond warmly and professionally.\n\nUser: " + prompt
        ))])
        return {"type": "text", "answer": resp.content}

    # ── Route 2: Structured (SQLite) — always try this first ─────────────────
    try:
        # Step 2a: Extract query filters
        filters = extract_search_filters(prompt, llm)
        
        # Step 2b: Query SQLite with extracted filters
        user_role = "admin" if is_user_admin else "user"
        matched_candidates = query_candidates_by_filters(filters, username, user_role)
        
        # Fall back to all candidates if query returned nothing, but limit to top 15
        if not matched_candidates:
            all_candidates = get_candidates_list(username, role=user_role)
            matched_candidates = all_candidates[:15]
        else:
            matched_candidates = matched_candidates[:15]
            
        if matched_candidates:
            for r in matched_candidates:
                for col in ['total_experience', 'pega_experience']:
                    if col in r:
                        try:
                            r[col] = float(r[col]) if str(r[col]).strip() != "" else 0
                        except:
                            r[col] = 0

            # Build compact summary of matches
            candidate_lines = []
            for r in matched_candidates:
                line = (
                    f"Name: {r.get('full_name','?')} | "
                    f"Total Exp: {r.get('total_experience',0)} yrs | "
                    f"Pega Exp: {r.get('pega_experience',0)} yrs | "
                    f"Skills: {r.get('skills','') or 'none'} | "
                    f"Certifications: {r.get('certifications','') or 'none'} | "
                    f"CTC: {r.get('ctc','') or '?'} | "
                    f"Notice: {r.get('notice_period','') or '?'} | "
                    f"Company: {r.get('current_organization','') or '?'} | "
                    f"Email: {r.get('email','') or '?'} | "
                    f"Phone: {r.get('phone','') or '?'}"
                )
                candidate_lines.append(line)
            candidates_text = "\n".join(candidate_lines)

            total_db_count = len(get_candidates_list(username, role=user_role))

            filter_prompt = f"""You are an expert HR data analyst. I have the following candidate database:

Total candidates in database: {total_db_count}

{candidates_text}

HR question: "{prompt}"

            Instructions:
            1. Read the question carefully.
            2. If the question asks to LIST / SHOW / FIND candidates:
               - Return EXACTLY this JSON format (no other text before or after):
               MATCH_RESULT: {{"type":"table","matched_names":["Candidate Name 1", "Candidate Name 2"],"intro":"Here are the candidates:"}}
            3. If the question asks for a COUNT or YES/NO:
               - Return EXACTLY: MATCH_RESULT: {{"type":"count","answer":"There are 3 candidates matching your criteria."}}
            4. If the question is about a SPECIFIC candidate's detail:
               - Return EXACTLY: MATCH_RESULT: {{"type":"table","matched_names":["Candidate Name"],"intro":"Here are the details:"}}
            5. IMPORTANT: Use ONLY exact full_name values from the database above. Do NOT invent names.
            6. 'Pega Exp: 0 yrs' = ZERO Pega experience.
            
            Respond with ONLY the MATCH_RESULT line, nothing else."""

            resp = llm.invoke([HumanMessage(content=filter_prompt)])
            ans  = resp.content.strip()

            if "MATCH_RESULT:" in ans:
                json_str = ans.split("MATCH_RESULT:")[1].strip()
                # Extract JSON object
                s, e = json_str.find('{'), json_str.rfind('}')
                if s != -1 and e != -1:
                    result = json.loads(json_str[s:e+1])

                    if result.get("type") == "count":
                        return {"type": "text", "answer": result.get("answer", ans)}

                    if result.get("type") == "table":
                        names = result.get("matched_names", [])
                        intro = result.get("intro", "Here are the matching candidates:")
                        if names:
                            matched_rows = _find_matching_rows(matched_candidates, names)
                            if matched_rows:
                                return {"type": "table", "answer": intro, "rows": matched_rows}
                        if matched_candidates:
                            fallback_rows = _find_matching_rows(matched_candidates, [r['full_name'] for r in matched_candidates[:5]])
                            if fallback_rows:
                                return {"type": "table", "answer": intro, "rows": fallback_rows}
                        return {"type": "text", "answer": "No candidates match your query. Try a different filter."}
    except Exception as e:
        print(f"Route 2 error, falling back to RAG: {e}")
        pass  # Fall through to RAG

    # ── Route 3: RAG (ChromaDB) — for very specific resume content ────────────
    if not os.path.exists(CHROMA_PATH):
        return {"type": "text", "answer": "No resumes uploaded yet. Please upload resumes first."}

    try:
        db        = Chroma(persist_directory=CHROMA_PATH, embedding_function=embeddings)
        retriever = db.as_retriever(search_kwargs={"k": 6})
        qa_prompt = PromptTemplate.from_template("""You are Hire AI, an expert HR recruitment assistant.
Answer using ONLY the resume context below. Be specific and structured.

Resume Context:
{context}

HR Question: {question}

Answer:""")

        def format_docs(docs):
            return "\n\n".join(d.page_content for d in docs)

        rag_chain = (
            {"context": retriever | format_docs, "question": RunnablePassthrough()}
            | qa_prompt
            | llm
            | StrOutputParser()
        )
        ans = rag_chain.invoke(prompt)
        return {"type": "text", "answer": ans}
    except Exception as e:
        return {"type": "text", "answer": f"Search error: {e}"}


# ── JD Matching ────────────────────────────────────────────────────────────────
class JDMatchRequest(BaseModel):
    job_description: str

@app.post("/api/match-jd")
def match_jd(req: JDMatchRequest):
    jd = req.job_description.strip()
    if not jd:
        raise HTTPException(status_code=400, detail="Empty Job Description")
    
    if not os.path.exists(CHROMA_PATH):
        return {"matches": []}

    embeddings, llm = get_models()
    db = Chroma(persist_directory=CHROMA_PATH, embedding_function=embeddings)
    
    try:
        docs = db.similarity_search(jd, k=15)
    except Exception:
        return {"matches": []}
    
    matched_sources = set()
    for d in docs:
        if 'source' in d.metadata:
            matched_sources.add(d.metadata['source'])
            
    if not matched_sources:
        return {"matches": []}

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    placeholders = ",".join("?" * len(matched_sources))
    cur.execute(f"SELECT * FROM candidate_metadata WHERE filename IN ({placeholders})", list(matched_sources))
    db_rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    if not is_user_admin:
        db_rows = [r for r in db_rows if r.get('created_by') and r['created_by'].lower() == username.lower()]

    if not db_rows:
        return {"matches": []}

    candidate_lines = []
    for r in db_rows:
        candidate_lines.append(
            f"Name: {r.get('full_name')} | "
            f"Total Experience: {r.get('total_experience')} yrs | "
            f"Pega Experience: {r.get('pega_experience')} yrs | "
            f"CDH Experience: {r.get('cdh_exp')} yrs | "
            f"Skills: {r.get('skills')} | "
            f"Certifications: {r.get('certifications')} | "
            f"Current Location: {r.get('current_location')} | "
            f"Preferred Locations: {r.get('pref_locations')}"
        )
    
    prompt = f"""You are an expert technical recruiter. Evaluate the following candidates against the Job Description "pin to pin".

Job Description:
{jd[:2000]}

Candidates to evaluate:
{chr(10).join(candidate_lines)}

Rules for evaluation:
1. Numeric Experience Matching: If a Job Description asks for "X+ years of experience", a candidate matches if their experience is greater than or equal to X.
   - For example: if Job Description requires "1+ years of experience in pega", then candidates with 3.0 years, 4.0 years, or 4.8 years of Pega experience all match perfectly because 3.0 >= 1.0, 4.0 >= 1.0, and 4.8 >= 1.0.
2. Certification Abbreviations:
   - CSSA is equivalent to any of: "PEGA Certified Senior System Architect", "Pega Certified Senior System Architect", "Certified Pega Senior System Architect", "Senior System Architect", or "CSSA".
   - CSA is equivalent to any of: "PEGA Certified System Architect", "Pega Certified System Architect", "Certified Pega System Architect", "System Architect", or "CSA".
   - LSA is equivalent to any of: "PEGA Certified Lead System Architect", "Pega Certified Lead System Architect", "Certified Pega Lead System Architect", "Lead System Architect", or "LSA".
3. Do not invent requirements. If the Job Description only mentions Pega experience, do NOT reject candidates for lacking CSSA or other unrelated certifications.
4. Location Matching: If the Job Description specifies a location requirement, a candidate matches if their Current Location or any of their Preferred Locations match the specified job location (e.g. if the Job Description mentions 'Chennai', a candidate with Current Location or Preferred Location 'Chennai' is a match).

For EACH candidate, decide if they match based on the rules. If they match, provide a 1-sentence explanation of why they are a good fit.
Format your response exactly as a JSON list of objects:
[
  {{
    "name": "Candidate Name",
    "reason": "1-sentence explanation of why they fit based on their specific experience, skills, and location"
  }}
]
Return ONLY the raw JSON block, no markdown, no other text."""

    try:
        resp = llm.invoke([HumanMessage(content=prompt)])
        raw = resp.content.strip()
        if "```json" in raw:
            raw = raw.split("```json")[1].split("```")[0].strip()
        elif "```" in raw:
            raw = raw.split("```")[1].split("```")[0].strip()
        
        start, end = raw.find('['), raw.rfind(']')
        if start != -1 and end != -1:
            raw = raw[start:end+1]
        
        ai_reasons = json.loads(raw)
        reason_map = {str(item.get("name", "")).strip().lower(): item.get("reason", "") for item in ai_reasons}
    except Exception:
        reason_map = {}

    matches = []
    for r in db_rows:
        name = str(r.get('full_name', '')).strip().lower()
        reason = reason_map.get(name, "Matched based on resume content similarity.")
        if reason == "Matched based on resume content similarity.":
            for k, v in reason_map.items():
                if k in name or name in k:
                    reason = v
                    break
        r['ai_reason'] = reason
        matches.append(r)
        
    return {"matches": matches}


# ── Jobs & JDs ─────────────────────────────────────────────────────────────────
class JobCreate(BaseModel):
    title: str
    description: str
    client_name: Optional[str] = ""
    contact_name: Optional[str] = ""
    client_phone: Optional[str] = ""
    account_manager: Optional[str] = ""
    assigned_recruiter: Optional[str] = ""
    target_date: Optional[str] = ""
    job_type: Optional[str] = ""
    job_status: Optional[str] = ""
    work_experience: Optional[str] = ""
    industry: Optional[str] = ""
    salary: Optional[str] = ""
    required_skills: Optional[str] = ""

class JobStatusUpdate(BaseModel):
    status: Optional[str] = None
    ai_reason: Optional[str] = None

class RegisterRequest(BaseModel):
    full_name: str
    username: str
    password: str
    email: str

class LoginRequest(BaseModel):
    username: str
    password: str

class UserPermissionsUpdate(BaseModel):
    is_hr: int
    is_admin: int
    is_external: Optional[int] = 0
    hidden_fields: Optional[str] = ""
    is_approved: Optional[int] = None

def has_digit(s: str) -> bool:
    return any(c.isdigit() for c in s)

@app.get("/api/jobs")
def list_jobs(request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        return []
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # Check if requesting user is external
    is_external = False
    is_user_admin = False
    if username:
        cur.execute("SELECT is_external, is_admin, role FROM users WHERE LOWER(username) = LOWER(?)", (username,))
        row = cur.fetchone()
        if row:
            is_external = (row['is_external'] == 1)
            is_user_admin = (row['is_admin'] == 1 or row['role'] == "admin" or is_admin_or_hr(username))
            
    if is_external:
        cur.execute("""
            SELECT j.* FROM jobs j
            JOIN job_shares js ON j.id = js.job_id
            WHERE LOWER(js.username) = LOWER(?)
            ORDER BY j.id DESC
        """, (username,))
    else:
        if is_user_admin:
            cur.execute("SELECT * FROM jobs ORDER BY id DESC")
        else:
            cur.execute("SELECT * FROM jobs WHERE LOWER(created_by) = LOWER(?) ORDER BY id DESC", (username,))
        
    jobs = [dict(r) for r in cur.fetchall()]
    for job in jobs:
        job_id = job['id']
        cur.execute("SELECT status, COUNT(*) as cnt FROM job_candidates WHERE job_id = ? GROUP BY status", (job_id,))
        counts = {r['status']: r['cnt'] for r in cur.fetchall()}
        job['matched_count'] = counts.get('matched', 0)
        job['selected_count'] = counts.get('selected', 0)
        
        # Get shared usernames
        cur.execute("SELECT username FROM job_shares WHERE job_id = ?", (job_id,))
        job['shared_with'] = [r['username'] for r in cur.fetchall()]
        
    conn.close()
    return jobs

@app.post("/api/jobs")
def create_job(job: JobCreate, request: Request):
    username = request.headers.get("x-user-username") or "admin"
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO jobs (
            title, description, client_name, contact_name, client_phone, account_manager,
            assigned_recruiter, target_date, job_type, job_status,
            work_experience, industry, salary, required_skills, created_by
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        job.title, job.description, job.client_name, job.contact_name, job.client_phone, job.account_manager,
        job.assigned_recruiter, job.target_date, job.job_type, job.job_status,
        job.work_experience, job.industry, job.salary, job.required_skills, username
    ))
    job_id = cur.lastrowid
    conn.commit()
    conn.close()
    
    try:
        match_candidates_for_job(job_id)
    except Exception as e:
        print(f"Error matching candidates for job {job_id}: {e}")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM jobs WHERE id = ?", (job_id,))
    updated_job = dict(cur.fetchone())
    
    cur.execute("SELECT status, COUNT(*) as cnt FROM job_candidates WHERE job_id = ? GROUP BY status", (job_id,))
    counts = {r['status']: r['cnt'] for r in cur.fetchall()}
    updated_job['matched_count'] = counts.get('matched', 0)
    updated_job['selected_count'] = counts.get('selected', 0)
    updated_job['shared_with'] = []
    username = request.headers.get("x-user-username")
    log_activity_db(username or "unknown", f"posted a Job Description for '{job.title}'")
    conn.close()
    
    return updated_job

@app.put("/api/jobs/{job_id}")
def update_job(job_id: int, job: JobCreate, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    created_by = row[0]
    
    if not is_admin_or_hr(username):
        if created_by and created_by.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")

    cur.execute("""
        UPDATE jobs SET 
            title = ?, description = ?, client_name = ?, contact_name = ?, client_phone = ?, account_manager = ?,
            assigned_recruiter = ?, target_date = ?, job_type = ?, job_status = ?,
            work_experience = ?, industry = ?, salary = ?, required_skills = ?
        WHERE id = ?
    """, (
        job.title, job.description, job.client_name, job.contact_name, job.client_phone, job.account_manager,
        job.assigned_recruiter, job.target_date, job.job_type, job.job_status,
        job.work_experience, job.industry, job.salary, job.required_skills,
        job_id
    ))
    if cur.rowcount == 0:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    conn.commit()
    conn.close()
    
    try:
        match_candidates_for_job(job_id)
    except Exception as e:
        print(f"Error matching candidates for job {job_id}: {e}")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM jobs WHERE id = ?", (job_id,))
    updated_job = dict(cur.fetchone())
    
    cur.execute("SELECT status, COUNT(*) as cnt FROM job_candidates WHERE job_id = ? GROUP BY status", (job_id,))
    counts = {r['status']: r['cnt'] for r in cur.fetchall()}
    updated_job['matched_count'] = counts.get('matched', 0)
    updated_job['selected_count'] = counts.get('selected', 0)
    
    cur.execute("SELECT username FROM job_shares WHERE job_id = ?", (job_id,))
    updated_job['shared_with'] = [r['username'] for r in cur.fetchall()]
    
    conn.close()
    
    return updated_job

@app.delete("/api/jobs/{job_id}")
def delete_job(job_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    cur.execute("SELECT title, created_by FROM jobs WHERE id = ?", (job_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    job_title, created_by = row
    
    if not is_admin_or_hr(username):
        if created_by and created_by.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")

    cur.execute("DELETE FROM job_candidates WHERE job_id = ?", (job_id,))
    cur.execute("DELETE FROM jobs WHERE id = ?", (job_id,))
    conn.commit()
    conn.close()
    log_activity_db(username or "unknown", f"deleted Job Description '{job_title}'")
    return {"message": "Job deleted"}

@app.get("/api/jobs/{job_id}/candidates")
def get_job_candidates(job_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Get job creator
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    job_row = cur.fetchone()
    if not job_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    job_creator = job_row['created_by']

    # Get requesting user's roles
    cur.execute("SELECT is_external, is_admin, role, full_name FROM users WHERE LOWER(username) = LOWER(?)", (username,))
    user_row = cur.fetchone()
    is_external = False
    is_user_admin = False
    user_full_name = ""
    if user_row:
        is_external = (user_row['is_external'] == 1)
        is_user_admin = (user_row['is_admin'] == 1 or user_row['role'] == "admin" or is_admin_or_hr(username))
        user_full_name = user_row['full_name']

    # Enforce permission checks:
    if not is_user_admin:
        if is_external:
            # Check share
            cur.execute("SELECT 1 FROM job_shares WHERE job_id = ? AND LOWER(username) = LOWER(?)", (job_id, username))
            if not cur.fetchone():
                conn.close()
                raise HTTPException(status_code=403, detail="Forbidden")
        else:
            # Internal user: must be creator of the job
            if job_creator and job_creator.lower() != username.lower():
                conn.close()
                raise HTTPException(status_code=403, detail="Forbidden")

    cur.execute("""
        SELECT c.*, jc.ai_reason, jc.status as job_status
        FROM candidate_metadata c
        JOIN job_candidates jc ON c.id = jc.candidate_id
        WHERE jc.job_id = ?
    """, (job_id,))
    candidates = [dict(row) for row in cur.fetchall()]
    conn.close()
    
    # Replace None values with empty string
    for row in candidates:
        for k, v in row.items():
            if v is None:
                row[k] = ""

    if is_external:
        # Mask all columns except for ID, name, AI reason, and status for data privacy
        allowed_keys = {'id', 'full_name', 'ai_reason', 'job_status', 'candidate_status'}
        for c in candidates:
            for key in list(c.keys()):
                if key not in allowed_keys:
                    c[key] = ""

    # Mask certifications for non-admin and non-HR users
    is_user_admin_or_hr = is_admin_or_hr(username)
    if not is_user_admin_or_hr:
        for row in candidates:
            row["certifications"] = "[HIDDEN]"

    if not is_admin_or_hr(username):
        keywords = get_masked_keywords()
        candidates = [mask_candidate_record(row, keywords) for row in candidates]
        
    candidates = apply_user_hidden_fields(candidates, username)
    return candidates

@app.get("/api/jobs/{job_id}/unmatched-candidates")
def get_unmatched_candidates(job_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Get job creator
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    job_row = cur.fetchone()
    if not job_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    job_creator = job_row['created_by']

    # Get requesting user's roles
    cur.execute("SELECT is_external, is_admin, role FROM users WHERE LOWER(username) = LOWER(?)", (username,))
    user_row = cur.fetchone()
    is_external = False
    is_user_admin = False
    if user_row:
        is_external = (user_row['is_external'] == 1)
        is_user_admin = (user_row['is_admin'] == 1 or user_row['role'] == "admin" or is_admin_or_hr(username))

    # Enforce permission checks:
    if not is_user_admin:
        if is_external:
            cur.execute("SELECT 1 FROM job_shares WHERE job_id = ? AND LOWER(username) = LOWER(?)", (job_id, username))
            if not cur.fetchone():
                conn.close()
                raise HTTPException(status_code=403, detail="Forbidden")
        else:
            if job_creator and job_creator.lower() != username.lower():
                conn.close()
                raise HTTPException(status_code=403, detail="Forbidden")

    # Filter unmatched candidates to only those owned by the user if not admin
    if is_user_admin:
        cur.execute("""
            SELECT * FROM candidate_metadata 
            WHERE id NOT IN (
                SELECT candidate_id FROM job_candidates WHERE job_id = ?
            )
            ORDER BY full_name ASC
        """, (job_id,))
    else:
        cur.execute("""
            SELECT * FROM candidate_metadata 
            WHERE LOWER(created_by) = LOWER(?)
            AND id NOT IN (
                SELECT candidate_id FROM job_candidates WHERE job_id = ?
            )
            ORDER BY full_name ASC
        """, (username, job_id))

    candidates = [dict(row) for row in cur.fetchall()]
    conn.close()

    # Replace None values with empty string
    for row in candidates:
        for k, v in row.items():
            if v is None:
                row[k] = ""

    username = request.headers.get("x-user-username")
    # Mask certifications for non-admin and non-HR users
    is_user_admin_or_hr = is_admin_or_hr(username)
    if not is_user_admin_or_hr:
        for row in candidates:
            row["certifications"] = "[HIDDEN]"

    if not is_admin_or_hr(username):
        keywords = get_masked_keywords()
        candidates = [mask_candidate_record(row, keywords) for row in candidates]
        
    candidates = apply_user_hidden_fields(candidates, username)
    return candidates

@app.post("/api/jobs/{job_id}/candidates/{candidate_id}")
def add_job_candidate(job_id: int, candidate_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    # Verify job creator and existence
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    job_row = cur.fetchone()
    if not job_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    job_creator = job_row[0]
    
    if not is_admin_or_hr(username):
        if job_creator and job_creator.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")
            
    # Check if already mapped
    cur.execute("SELECT 1 FROM job_candidates WHERE job_id = ? AND candidate_id = ?", (job_id, candidate_id))
    if cur.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Candidate is already matched to this job.")
        
    cur.execute("""
        INSERT INTO job_candidates (job_id, candidate_id, ai_reason, status)
        VALUES (?, ?, 'Manually associated by recruiter.', 'matched')
    """, (job_id, candidate_id))
    
    # Fetch candidate name for logging
    cur.execute("SELECT full_name FROM candidate_metadata WHERE id = ?", (candidate_id,))
    cand_row = cur.fetchone()
    cand_name = cand_row[0] if cand_row else f"ID {candidate_id}"
    
    # Fetch job title for logging
    cur.execute("SELECT title FROM jobs WHERE id = ?", (job_id,))
    job_row = cur.fetchone()
    job_title = job_row[0] if job_row else f"ID {job_id}"
    
    conn.commit()
    conn.close()
    
    # Try to log activity
    try:
        log_activity_db("recruiter", f"manually matched candidate '{cand_name}' to job '{job_title}'")
    except Exception as e:
        print("Failed to log activity:", e)
        
    return {"message": "Candidate associated with job successfully"}

@app.put("/api/jobs/{job_id}/candidates/{candidate_id}")
def update_job_candidate_status(job_id: int, candidate_id: int, update: JobStatusUpdate, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    # Verify job creator and existence
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    job_row = cur.fetchone()
    if not job_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    job_creator = job_row[0]
    
    if not is_admin_or_hr(username):
        if job_creator and job_creator.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")
    
    if update.status is not None:
        if update.status not in ["matched", "selected"]:
            conn.close()
            raise HTTPException(status_code=400, detail="Invalid status")
        cur.execute("UPDATE job_candidates SET status = ? WHERE job_id = ? AND candidate_id = ?", (update.status, job_id, candidate_id))
        
    if update.ai_reason is not None:
        cur.execute("UPDATE job_candidates SET ai_reason = ? WHERE job_id = ? AND candidate_id = ?", (update.ai_reason, job_id, candidate_id))
        
    conn.commit()
    conn.close()
    return {"message": "Status updated"}

@app.delete("/api/jobs/{job_id}/candidates/{candidate_id}")
def delete_job_candidate(job_id: int, candidate_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)

    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    # Verify job creator and existence
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    job_row = cur.fetchone()
    if not job_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    job_creator = job_row[0]
    
    if not is_admin_or_hr(username):
        if job_creator and job_creator.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")
    cur.execute("DELETE FROM job_candidates WHERE job_id = ? AND candidate_id = ?", (job_id, candidate_id))
    conn.commit()
    conn.close()
    return {"message": "Candidate removed from job"}

def match_candidates_for_job(job_id: int):
    # This endpoint finds matching candidates for a job and saves them to job_candidates
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT description, created_by FROM jobs WHERE id = ?", (job_id,))
    job = cur.fetchone()
    if not job:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    
    jd = job['description']
    job_creator = job['created_by']
    
    # Query candidates directly from the database (only those owned by job creator unless admin or HR)
    if job_creator and not is_admin_or_hr(job_creator):
        cur.execute("SELECT * FROM candidate_metadata WHERE LOWER(created_by) = LOWER(?)", (job_creator,))
    else:
        cur.execute("SELECT * FROM candidate_metadata")
    db_rows = [dict(r) for r in cur.fetchall()]

    if not db_rows:
        conn.close()
        return {"message": "No candidates found in DB"}

    candidate_lines = []
    for r in db_rows:
        candidate_lines.append(
            f"Name: {r.get('full_name')} | "
            f"Total Experience: {r.get('total_experience')} yrs | "
            f"Pega Experience: {r.get('pega_experience')} yrs | "
            f"CDH Experience: {r.get('cdh_exp')} yrs | "
            f"Skills: {r.get('skills')} | "
            f"Certifications: {r.get('certifications')} | "
            f"Current Location: {r.get('current_location')} | "
            f"Preferred Locations: {r.get('pref_locations')}"
        )
    
    emb, llm = get_models()
    
    prompt = f"""You are an expert technical recruiter. Evaluate these candidates against the Job Description "pin to pin".

Job Description:
{jd[:2000]}

Candidates to evaluate:
{chr(10).join(candidate_lines)}

Rules for evaluation:
1. Numeric Experience Matching: If a Job Description asks for "X+ years of experience", a candidate matches if their experience is greater than or equal to X.
   - For example: if Job Description requires "1+ years of experience in pega", then candidates with 3.0 years, 4.0 years, or 4.8 years of Pega experience all match perfectly because 3.0 >= 1.0, 4.0 >= 1.0, and 4.8 >= 1.0.
2. Certification Abbreviations:
   - CSSA is equivalent to any of: "PEGA Certified Senior System Architect", "Pega Certified Senior System Architect", "Certified Pega Senior System Architect", "Senior System Architect", or "CSSA".
   - CSA is equivalent to any of: "PEGA Certified System Architect", "Pega Certified System Architect", "Certified Pega System Architect", "System Architect", or "CSA".
   - LSA is equivalent to any of: "PEGA Certified Lead System Architect", "Pega Certified Lead System Architect", "Certified Pega Lead System Architect", "Lead System Architect", or "LSA".
3. Do not invent requirements. If the Job Description only mentions Pega experience, do NOT reject candidates for lacking CSSA or other unrelated certifications.
4. Location Matching: If the Job Description specifies a location requirement, a candidate matches if their Current Location or any of their Preferred Locations match the specified job location (e.g. if the Job Description mentions 'Chennai', a candidate with Current Location or Preferred Location 'Chennai' is a match).

Format your response exactly as a JSON list of objects for matching candidates only:
[
  {{
    "name": "Candidate Name",
    "reason": "1-sentence explanation of why they fit based on their specific experience, skills, and location"
  }}
]
Return ONLY the raw JSON block, no markdown, no other text."""

    try:
        resp = llm.invoke([HumanMessage(content=prompt)])
        raw = resp.content.strip()
        if "```json" in raw: raw = raw.split("```json")[1].split("```")[0].strip()
        elif "```" in raw: raw = raw.split("```")[1].split("```")[0].strip()
        start, end = raw.find('['), raw.rfind(']')
        if start != -1 and end != -1: raw = raw[start:end+1]
        ai_reasons = json.loads(raw)
        reason_map = {str(item.get("name", "")).strip().lower(): item.get("reason", "") for item in ai_reasons}
    except Exception:
        reason_map = {}

    # Clear old automatic matches for this job that haven't been selected
    cur.execute("DELETE FROM job_candidates WHERE job_id = ? AND status = 'matched'", (job_id,))

    matches_added = 0
    for r in db_rows:
        name = str(r.get('full_name', '')).strip().lower()
        reason = None
        for k, v in reason_map.items():
            if k in name or name in k:
                reason = v
                break
        
        if reason:
            # Upsert into job_candidates
            cur.execute("""
                INSERT INTO job_candidates (job_id, candidate_id, ai_reason, status) 
                VALUES (?, ?, ?, 'matched')
                ON CONFLICT(job_id, candidate_id) DO UPDATE SET ai_reason = excluded.ai_reason
            """, (job_id, r['id'], reason))
            
            # Also mark candidate as qualified
            cur.execute("UPDATE candidate_metadata SET is_qualified = 1 WHERE id = ?", (r['id'],))
            matches_added += 1

    conn.commit()
    conn.close()
    return {"message": f"Successfully matched and added {matches_added} candidates to job"}

@app.post("/api/jobs/{job_id}/match")
def match_candidates_for_job_endpoint(job_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    role = get_user_role(username)
    
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
    created_by = row[0]
    conn.close()
    
    if not is_admin_or_hr(username):
        if created_by and created_by.lower() != username.lower():
            raise HTTPException(status_code=403, detail="Forbidden")
            
    return match_candidates_for_job(job_id)

# ── Reset ──────────────────────────────────────────────────────────────────────
@app.post("/api/reset")
def reset_all(request: Request):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.execute("DELETE FROM candidate_metadata")
    conn.execute("DELETE FROM change_requests")
    conn.execute("DELETE FROM activity_logs")
    conn.commit()
    conn.close()
    try:
        if os.path.exists(CHROMA_PATH):
            shutil.rmtree(CHROMA_PATH)
    except Exception:
        pass
    return {"status": "reset complete"}

# ── Auth & Admin Endpoints ───────────────────────────────────────────────────────
@app.post("/api/auth/register")
def register(req: RegisterRequest):
    if not has_digit(req.password):
        raise HTTPException(status_code=400, detail="Password must contain at least one digit")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    username_exists = False
    email_limit_exceeded = False
    try:
        cur.execute("SELECT 1 FROM users WHERE LOWER(username) = LOWER(?)", (req.username,))
        if cur.fetchone():
            username_exists = True
        else:
            base_email = get_base_email(req.email)
            cur.execute("SELECT email FROM users")
            all_emails = [r[0] for r in cur.fetchall() if r[0]]
            same_email_count = sum(1 for e in all_emails if get_base_email(e) == base_email)
            if same_email_count >= 5:
                email_limit_exceeded = True
            else:
                is_approved_val = 1 if req.username.lower() in ("admin", "user", "boopathi", "praveen", "harish", "sabari") else 0
                cur.execute("INSERT INTO users (full_name, username, password, email, role, is_approved) VALUES (?, ?, ?, ?, 'user', ?)",
                            (req.full_name, req.username, req.password, req.email, is_approved_val))
                if is_approved_val == 0:
                    cur.execute("""
                        INSERT INTO change_requests (username, action_type, target_id, payload, description, status)
                        VALUES (?, 'approve_user', ?, NULL, ?, 'pending')
                    """, (req.username, req.username, f"Approve access request for registered user {req.full_name} (@{req.username})"))
                conn.commit()
    except sqlite3.IntegrityError:
        username_exists = True
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
    conn.close()
    
    if username_exists:
        raise HTTPException(status_code=400, detail="Username already exists")
    if email_limit_exceeded:
        raise HTTPException(status_code=400, detail="Maximum of 5 accounts can be created with the same email address.")
        
    return {"status": "registered", "username": req.username}

@app.post("/api/auth/login")
def login(req: LoginRequest):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(?) AND password = ?", (req.username, req.password))
    user = cur.fetchone()
    conn.close()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    user_dict = dict(user)
    if user_dict.get("is_approved", 0) == 0:
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    return {
        "username": user_dict["username"],
        "full_name": user_dict["full_name"],
        "role": user_dict["role"],
        "is_hr": user_dict.get("is_hr", 0),
        "is_admin": user_dict.get("is_admin", 0),
        "is_external": user_dict.get("is_external", 0),
        "is_approved": user_dict.get("is_approved", 0),
        "email": user_dict.get("email", ""),
        "hidden_fields": user_dict.get("hidden_fields", "")
    }

class FirebaseSyncRequest(BaseModel):
    email: str
    full_name: str
    username: str

@app.post("/api/auth/firebase-sync")
def firebase_sync(req: FirebaseSyncRequest):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    cur.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(?)", (req.username,))
    user = cur.fetchone()
    
    if not user:
        # Check email limit
        base_email = get_base_email(req.email)
        cur.execute("SELECT email FROM users")
        all_emails = [r[0] for r in cur.fetchall() if r[0]]
        same_email_count = sum(1 for e in all_emails if get_base_email(e) == base_email)
        if same_email_count >= 5:
            conn.close()
            raise HTTPException(status_code=400, detail="Maximum of 5 accounts can be created with the same email address.")
            
        try:
            is_approved_val = 1 if req.username.lower() in ("admin", "user", "boopathi", "praveen", "harish", "sabari") else 0
            cur.execute("INSERT INTO users (full_name, username, password, email, role, is_approved) VALUES (?, ?, ?, ?, 'user', ?)",
                        (req.full_name, req.username, "firebase_auth_managed", req.email, is_approved_val))
            if is_approved_val == 0:
                cur.execute("""
                    INSERT INTO change_requests (username, action_type, target_id, payload, description, status)
                    VALUES (?, 'approve_user', ?, NULL, ?, 'pending')
                """, (req.username, req.username, f"Approve access request for registered user {req.full_name} (@{req.username})"))
            conn.commit()
            
            cur.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(?)", (req.username,))
            user = cur.fetchone()
            log_activity_db(req.username, "registered an account via Firebase")
        except Exception as e:
            conn.close()
            raise HTTPException(status_code=500, detail=f"Database synchronization error: {str(e)}")
    else:
        # User exists, let's make sure email is synced if it is missing
        try:
            user_dict = dict(user)
            if (not user_dict.get("email") or user_dict.get("email") == "") and req.email:
                cur.execute("UPDATE users SET email = ? WHERE LOWER(username) = LOWER(?)", (req.email, req.username.lower()))
                conn.commit()
                # Fetch again to get updated user dict
                cur.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(?)", (req.username,))
                user = cur.fetchone()
        except Exception as e:
            print(f"Failed to update user email during firebase sync: {e}")
            
    conn.close()
    user_dict = dict(user)
    if user_dict.get("is_approved", 0) == 0:
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    return {
        "username": user_dict["username"],
        "full_name": user_dict["full_name"],
        "role": user_dict["role"],
        "is_hr": user_dict.get("is_hr", 0),
        "is_admin": user_dict.get("is_admin", 0),
        "is_external": user_dict.get("is_external", 0),
        "is_approved": user_dict.get("is_approved", 0),
        "email": user_dict.get("email", ""),
        "hidden_fields": user_dict.get("hidden_fields", "")
    }

@app.get("/api/auth/check-exists")
def check_user_exists(username: str, email: str):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    # 1. Check if username exists
    cur.execute("SELECT 1 FROM users WHERE LOWER(username) = LOWER(?)", (username,))
    if cur.fetchone():
        conn.close()
        return {"exists": True, "reason": "Username already exists."}
        
    # 2. Check email limit (5 accounts)
    base_email = get_base_email(email)
    cur.execute("SELECT email FROM users")
    all_emails = [r[0] for r in cur.fetchall() if r[0]]
    same_email_count = sum(1 for e in all_emails if get_base_email(e) == base_email)
    if same_email_count >= 5:
        conn.close()
        return {"exists": True, "reason": "Maximum of 5 accounts can be created with the same email address."}
        
    conn.close()
    return {"exists": False}

OTP_STORE = {}

@app.get("/api/auth/get-email")
def get_email(username: str):
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT email FROM users WHERE LOWER(username) = LOWER(?)", (username,))
    row = cur.fetchone()
    conn.close()
    if row and row[0]:
        return {"email": row[0]}
    # Fallback
    return {"email": f"{username.lower()}@hireai.local"}

@app.get("/api/auth/status")
def get_user_status(request: Request):
    username = request.headers.get("x-user-username")
    if not username:
        raise HTTPException(status_code=401, detail="Not authenticated")
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE LOWER(username) = LOWER(?)", (username,))
    row = cur.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="User not found")
    user_dict = dict(row)
    return {
        "username": user_dict["username"],
        "full_name": user_dict["full_name"],
        "role": user_dict["role"],
        "is_hr": user_dict.get("is_hr", 0),
        "is_admin": user_dict.get("is_admin", 0),
        "is_external": user_dict.get("is_external", 0),
        "is_approved": user_dict.get("is_approved", 0),
        "email": user_dict.get("email", ""),
        "hidden_fields": user_dict.get("hidden_fields", "")
    }

class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    email: str
    otp: str
    new_password: str

def send_otp_email(to_email: str, otp: str, raise_on_error: bool = False) -> bool:
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_sender = os.getenv("SMTP_SENDER", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    
    if not smtp_sender or not smtp_password:
        print(f"[SMTP] Credentials not set. Skipped sending real email.")
        if raise_on_error:
            raise ValueError("SMTP email credentials are not configured in your .env file. Please set SMTP_SENDER and SMTP_PASSWORD.")
        return False
        
    subject = "Hire AI - Password Reset OTP"
    body = f"Hi,\n\nYour 6-digit OTP code to reset your Hire AI password is: {otp}\n\nThis code is valid for 5 minutes.\n\nBest regards,\nHire AI Team"
    
    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_sender
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=10.0)
        server.starttls()
        server.login(smtp_sender, smtp_password)
        server.sendmail(smtp_sender, to_email, msg.as_string())
        server.quit()
        print(f"[SMTP] Successfully sent OTP email to {to_email}")
        return True
    except Exception as e:
        print(f"[SMTP] Failed to send email to {to_email}: {str(e)}")
        if raise_on_error:
            raise RuntimeError(f"Failed to send email: {str(e)}")
        return False

@app.post("/api/auth/forgot-password/request")
def request_otp(req: ForgotPasswordRequest):
    import random
    import time
    email = req.email.strip().lower()
    base_email = get_base_email(email)
    
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT email, username FROM users")
    users = cur.fetchall()
    matching_users = [u for u in users if u[0] and get_base_email(u[0]) == base_email]
    
    if not matching_users:
        # If no user matches the email directly, attempt heuristic match on username using the local part of email.
        # This helps seeded/default or manually created users whose email is NULL or empty in SQLite.
        if "@" in email:
            local_part = email.split("@")[0].split("+")[0].strip().lower()
            # Try to find a user where the username matches or starts/ends with the local part
            matching_username = None
            for u_email, u_name in users:
                u_name_lower = u_name.strip().lower()
                if u_name_lower == local_part or u_name_lower.startswith(local_part) or local_part.startswith(u_name_lower):
                    matching_username = u_name
                    break
            
            if matching_username:
                try:
                    cur.execute("UPDATE users SET email = ? WHERE LOWER(username) = LOWER(?)", (email, matching_username))
                    conn.commit()
                    matching_users = [(email, matching_username)]
                    print(f"[AUTH] Updated user {matching_username} email to {email} via heuristic match")
                except Exception as e:
                    print(f"[AUTH] Failed to update user email during heuristic match: {e}")
    conn.close()
    
    if not matching_users:
        raise HTTPException(status_code=404, detail="No registered account found with this email address.")
        
    otp = f"{random.randint(100000, 999999)}"
    OTP_STORE[email] = {
        "otp": otp,
        "expires_at": time.time() + 300.0
    }
    
    print(f"========================================")
    print(f"[OTP SIMULATION] Password reset OTP for {email}: {otp}")
    print(f"========================================")
    
    require_real = os.getenv("REQUIRE_REAL_EMAIL", "true").lower() == "true"
    
    sent = False
    try:
        sent = send_otp_email(email, otp, raise_on_error=require_real)
    except Exception as e:
        print(f"[SMTP] Error during send_otp_email: {str(e)}")
        # If we require real email and it fails, but SMTP credentials are NOT configured,
        # we treat it as a warning and fall back to returning simulated OTP rather than raising a 500 error.
        smtp_sender = os.getenv("SMTP_SENDER", "")
        smtp_password = os.getenv("SMTP_PASSWORD", "")
        if require_real and smtp_sender and smtp_password:
            # Only raise 500 if the credentials are set but the actual mail delivery fails
            if email in OTP_STORE:
                del OTP_STORE[email]
            raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")
        else:
            # Fall back to simulated OTP display
            sent = False
            
    return {
        "status": "success",
        "otp": otp if not sent else None,
        "sent_real_email": sent,
        "message": "OTP has been sent to your email." if sent else "OTP has been simulated. Please use the demo code shown."
    }

@app.post("/api/auth/forgot-password/reset")
def reset_password(req: ResetPasswordRequest):
    import time
    email = req.email.strip().lower()
    otp_code = req.otp.strip()
    new_pass = req.new_password
    
    if not has_digit(new_pass):
        raise HTTPException(status_code=400, detail="Password must contain at least one digit")
        
    if email not in OTP_STORE:
        raise HTTPException(status_code=400, detail="No active password reset request found for this email.")
        
    stored = OTP_STORE[email]
    if time.time() > stored["expires_at"]:
        del OTP_STORE[email]
        raise HTTPException(status_code=400, detail="OTP has expired. Please request a new one.")
        
    if stored["otp"] != otp_code:
        raise HTTPException(status_code=400, detail="Invalid OTP code. Please try again.")
        
    del OTP_STORE[email]
    
    base_email = get_base_email(email)
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT email, username FROM users")
    users = cur.fetchall()
    matching_users = [u for u in users if get_base_email(u[0]) == base_email]
    
    for mu in matching_users:
        cur.execute("UPDATE users SET password = ? WHERE LOWER(email) = LOWER(?)", (new_pass, mu[0].lower()))
        
    conn.commit()
    conn.close()
    
    return {
        "status": "success",
        "message": "Password reset successfully. You can now log in with your new password."
    }

@app.post("/api/candidates")
def add_candidate_manually(request: Request, body: dict):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.pragma("table_info(candidate_metadata)")
    # wait, cur.execute("PRAGMA table_info(candidate_metadata)")
    cur.execute("PRAGMA table_info(candidate_metadata)")
    valid_cols = [c[1] for c in cur.fetchall()]
    
    insert_data = {}
    for col in valid_cols:
        if col in ('id', 'timestamp'):
            continue
        if col in body:
            insert_data[col] = body[col]
            
    if 'source' not in insert_data or not insert_data['source']:
        insert_data['source'] = 'Manual Entry'
    if 'candidate_status' not in insert_data or not insert_data['candidate_status']:
        insert_data['candidate_status'] = 'New'
    insert_data['created_by'] = username or "admin"
        
    if not insert_data.get('full_name'):
        conn.close()
        raise HTTPException(status_code=400, detail="Candidate Name is required")
        
    cols_str = ", ".join(insert_data.keys())
    placeholders = ", ".join(["?"] * len(insert_data))
    vals = list(insert_data.values())
    
    try:
        cur.execute(f"INSERT INTO candidate_metadata ({cols_str}) VALUES ({placeholders})", vals)
        conn.commit()
        log_activity_db(username or "unknown", f"manually added candidate '{insert_data.get('full_name')}'")
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
        
    conn.close()
    return {"status": "success", "message": "Candidate added successfully"}

@app.get("/api/admin/requests")
def list_change_requests(request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM change_requests ORDER BY created_at DESC")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.get("/api/admin/users")
def list_users(request: Request):
    username = request.headers.get("x-user-username")
    if not is_admin_or_hr(username):
        raise HTTPException(status_code=403, detail="Forbidden")
        
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT id, full_name, username, role, is_hr, is_admin, is_external, hidden_fields, is_approved, email FROM users")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def is_admin_or_hr(username: str) -> bool:
    if not username:
        return False
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT is_admin, is_hr FROM users WHERE LOWER(username) = LOWER(?)", (username,))
    row = cur.fetchone()
    conn.close()
    if row:
        is_admin, is_hr = row
        return is_admin == 1 or is_hr == 1
    return False

def get_user_hidden_fields(username: str) -> list[str]:
    if not username:
        return []
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    try:
        cur.execute("SELECT hidden_fields FROM users WHERE LOWER(username) = LOWER(?)", (username,))
        row = cur.fetchone()
        if row and row[0]:
            conn.close()
            return [f.strip().lower() for f in row[0].split(",") if f.strip()]
    except Exception as e:
        pass
    conn.close()
    return []

def apply_user_hidden_fields(rows: list[dict], username: str) -> list[dict]:
    hidden = get_user_hidden_fields(username)
    if not hidden:
        return rows
    for r in rows:
        for field in hidden:
            if field in r:
                r[field] = "[HIDDEN]"
    return rows

@app.put("/api/admin/users/{user_id}/permissions")
def update_user_permissions_endpoint(user_id: int, body: UserPermissionsUpdate, request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    is_external_val = body.is_external if body.is_external is not None else 0
    is_hr_val = body.is_hr
    is_admin_val = body.is_admin
    hidden_fields_val = body.hidden_fields if body.hidden_fields is not None else ""
    
    if is_external_val == 1:
        is_hr_val = 0
        is_admin_val = 0
        
    new_role = "admin" if is_admin_val == 1 else "user"
    if body.is_approved is not None:
        cur.execute("UPDATE users SET is_hr = ?, is_admin = ?, is_external = ?, role = ?, hidden_fields = ?, is_approved = ? WHERE id = ?", 
                    (is_hr_val, is_admin_val, is_external_val, new_role, hidden_fields_val, body.is_approved, user_id))
    else:
        cur.execute("UPDATE users SET is_hr = ?, is_admin = ?, is_external = ?, role = ?, hidden_fields = ? WHERE id = ?", 
                    (is_hr_val, is_admin_val, is_external_val, new_role, hidden_fields_val, user_id))
    conn.commit()
    conn.close()
    return {"status": "updated"}

class JobShareRequest(BaseModel):
    usernames: list[str]

@app.post("/api/jobs/{job_id}/share")
def share_job(job_id: int, req: JobShareRequest, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    if not is_admin_or_hr(username):
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    # Verify job creator and existence
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    job_row = cur.fetchone()
    if not job_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
        
    created_by = job_row[0]
    role = get_user_role(username)
    if role != "admin":
        if created_by and created_by.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")
    
    # Verify job exists
    cur.execute("SELECT 1 FROM jobs WHERE id = ?", (job_id,))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
        
    # Clear existing shares
    cur.execute("DELETE FROM job_shares WHERE job_id = ?", (job_id,))
    
    # Insert new shares
    for u in req.usernames:
        cur.execute("INSERT INTO job_shares (job_id, username) VALUES (?, ?)", (job_id, u))
        
    conn.commit()
    conn.close()
    
    # Log activity
    job_title = get_job_title(job_id)
    log_activity_db(username or "unknown", f"shared Job Description '{job_title}' with {len(req.usernames)} external users")
    
    return {"status": "shared"}

@app.get("/api/jobs/{job_id}/shares")
def get_job_shares(job_id: int, request: Request):
    username = request.headers.get("x-user-username")
    if not is_user_approved(username):
        raise HTTPException(status_code=403, detail="Access denied. Your account is pending admin approval.")
    if not is_admin_or_hr(username):
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    # Verify job creator and existence
    cur.execute("SELECT created_by FROM jobs WHERE id = ?", (job_id,))
    job_row = cur.fetchone()
    if not job_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Job not found")
        
    created_by = job_row[0]
    role = get_user_role(username)
    if role != "admin":
        if created_by and created_by.lower() != username.lower():
            conn.close()
            raise HTTPException(status_code=403, detail="Forbidden")
    cur.execute("SELECT username FROM job_shares WHERE job_id = ?", (job_id,))
    rows = cur.fetchall()
    conn.close()
    
    return [r[0] for r in rows]

class MaskedKeywordCreate(BaseModel):
    keyword: str

@app.get("/api/admin/masked-keywords")
def get_admin_masked_keywords(request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
    return get_masked_keywords()

@app.post("/api/admin/masked-keywords")
def add_admin_masked_keyword(req: MaskedKeywordCreate, request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
    kw = req.keyword.strip()
    if not kw:
        raise HTTPException(status_code=400, detail="Keyword cannot be empty")
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    try:
        cur.execute("INSERT OR IGNORE INTO masked_keywords (keyword) VALUES (?)", (kw,))
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
    conn.close()
    log_activity_db(username, f"added masked keyword '{kw}'")
    return {"status": "added", "keyword": kw}

@app.delete("/api/admin/masked-keywords/{keyword}")
def delete_admin_masked_keyword(keyword: str, request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
    kw = keyword.strip()
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM masked_keywords WHERE keyword = ?", (kw,))
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
    conn.close()
    log_activity_db(username, f"deleted masked keyword '{kw}'")
    return {"status": "deleted", "keyword": kw}

@app.delete("/api/admin/users/{user_id}")
def delete_user_endpoint(user_id: int, request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    # Prevent self-deletion and get email/fullname
    cur.execute("SELECT username, email, full_name FROM users WHERE id = ?", (user_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")
        
    deleted_username = row[0]
    deleted_email = row[1]
    deleted_fullname = row[2]
    
    if deleted_username.lower() == username.lower():
        conn.close()
        raise HTTPException(status_code=400, detail="You cannot delete yourself.")
        
    # Delete from users table
    cur.execute("DELETE FROM users WHERE id = ?", (user_id,))
    
    # Delete from change_requests (where requester or target is this user)
    cur.execute("DELETE FROM change_requests WHERE LOWER(username) = LOWER(?) OR LOWER(target_id) = LOWER(?)", (deleted_username, deleted_username))
    
    # Delete from job_shares
    cur.execute("DELETE FROM job_shares WHERE LOWER(username) = LOWER(?)", (deleted_username,))
    
    # Delete from team_members (by full name and by username, if exists)
    cur.execute("DELETE FROM team_members WHERE LOWER(name) = LOWER(?) OR LOWER(name) = LOWER(?)", (deleted_fullname, deleted_username))
    
    # Delete from activity_logs
    cur.execute("DELETE FROM activity_logs WHERE LOWER(username) = LOWER(?)", (deleted_username,))
    
    # Delete candidate records and their resume files owned/created by this user
    cur.execute("SELECT id, filename FROM candidate_metadata WHERE LOWER(created_by) = LOWER(?)", (deleted_username,))
    candidates = cur.fetchall()
    if candidates:
        candidate_ids = [c[0] for c in candidates]
        placeholders = ",".join("?" for _ in candidate_ids)
        
        # Delete resume files from disk
        for c_id, fname in candidates:
            if fname:
                fpath = os.path.join(UPLOAD_DIR, fname)
                if os.path.exists(fpath):
                    try:
                        os.remove(fpath)
                    except Exception as e:
                        print(f"Error removing resume file {fpath}: {e}")
                        
        # Delete from DB
        cur.execute(f"DELETE FROM job_candidates WHERE candidate_id IN ({placeholders})", candidate_ids)
        cur.execute(f"DELETE FROM candidate_metadata WHERE id IN ({placeholders})", candidate_ids)
        
    # Delete job records created by this user
    cur.execute("SELECT id FROM jobs WHERE LOWER(created_by) = LOWER(?)", (deleted_username,))
    job_ids = [r[0] for r in cur.fetchall()]
    if job_ids:
        placeholders = ",".join("?" for _ in job_ids)
        cur.execute(f"DELETE FROM job_candidates WHERE job_id IN ({placeholders})", job_ids)
        cur.execute(f"DELETE FROM job_shares WHERE job_id IN ({placeholders})", job_ids)
        cur.execute(f"DELETE FROM jobs WHERE id IN ({placeholders})", job_ids)
        
    conn.commit()
    conn.close()
    
    log_activity_db(username, f"completely deleted user '{deleted_username}' from system")
    return {"status": "deleted"}

@app.post("/api/admin/requests/{request_id}/approve")
def approve_change_request(request_id: int, request: Request, background_tasks: BackgroundTasks):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM change_requests WHERE id = ?", (request_id,))
    req_row = cur.fetchone()
    if not req_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Request not found")
        
    req_data = dict(req_row)
    if req_data["status"] != "pending":
        conn.close()
        raise HTTPException(status_code=400, detail="Request is already resolved")
        
    action_type = req_data["action_type"]
    target_id = req_data["target_id"]
    payload = req_data["payload"]
    
    try:
        if action_type == "add_column":
            col_data = json.loads(payload)
            clean_key = re.sub(r'[^a-zA-Z0-9_]', '', col_data["col_key"].replace(' ', '_')).lower()
            cur.execute("PRAGMA table_info(candidate_metadata)")
            existing = [c[1] for c in cur.fetchall()]
            if clean_key not in existing:
                cur.execute(f"ALTER TABLE candidate_metadata ADD COLUMN {clean_key} TEXT")
                cur.execute("INSERT OR IGNORE INTO custom_columns (col_key, col_label, description) VALUES (?, ?, ?)", 
                            (clean_key, col_data["col_label"], col_data["description"]))
                
        elif action_type == "delete_column":
            col_key = target_id
            cur.execute("DELETE FROM custom_columns WHERE col_key=?", (col_key,))
            try:
                cur.execute(f"ALTER TABLE candidate_metadata DROP COLUMN {col_key}")
            except Exception:
                pass
                
        elif action_type == "update_candidate":
            candidate_id = int(target_id)
            updates = json.loads(payload)
            set_clause = ", ".join(f"{k}=?" for k in updates)
            cur.execute(
                f"UPDATE candidate_metadata SET {set_clause} WHERE id=?",
                list(updates.values()) + [candidate_id]
            )
            match_related_fields = {
                'full_name', 'total_experience', 'pega_experience', 'cdh_exp',
                'skills', 'certifications', 'current_location', 'pref_locations'
            }
            if any(field in updates for field in match_related_fields):
                background_tasks.add_task(match_candidate_to_all_jobs, candidate_id)
                
        elif action_type == "delete_candidate":
            candidate_id = int(target_id)
            cur.execute("DELETE FROM job_candidates WHERE candidate_id=?", (candidate_id,))
            cur.execute("DELETE FROM candidate_metadata WHERE id=?", (candidate_id,))
            
        elif action_type == "approve_resume":
            candidate_id = int(target_id)
            cur.execute("UPDATE candidate_metadata SET is_approved = 1 WHERE id = ?", (candidate_id,))
            background_tasks.add_task(match_candidate_to_all_jobs, candidate_id)
            
        elif action_type == "create_job":
            job_data = json.loads(payload)
            cur.execute("INSERT INTO jobs (title, description) VALUES (?, ?)", (job_data["title"], job_data["description"]))
            job_id = cur.lastrowid
            background_tasks.add_task(match_candidates_for_job, job_id)
            
        elif action_type == "update_job":
            job_id = int(target_id)
            job_data = json.loads(payload)
            cur.execute("UPDATE jobs SET title = ?, description = ? WHERE id = ?", (job_data["title"], job_data["description"], job_id))
            background_tasks.add_task(match_candidates_for_job, job_id)
            
        elif action_type == "delete_job":
            job_id = int(target_id)
            cur.execute("DELETE FROM job_candidates WHERE job_id = ?", (job_id,))
            cur.execute("DELETE FROM jobs WHERE id = ?", (job_id,))
            
        elif action_type == "update_job_candidate":
            job_id, candidate_id = map(int, target_id.split(":"))
            update_data = json.loads(payload)
            if "status" in update_data and update_data["status"] is not None:
                cur.execute("UPDATE job_candidates SET status = ? WHERE job_id = ? AND candidate_id = ?", 
                            (update_data["status"], job_id, candidate_id))
            if "ai_reason" in update_data and update_data["ai_reason"] is not None:
                cur.execute("UPDATE job_candidates SET ai_reason = ? WHERE job_id = ? AND candidate_id = ?", 
                            (update_data["ai_reason"], job_id, candidate_id))
                                
        elif action_type == "delete_job_candidate":
            job_id, candidate_id = map(int, target_id.split(":"))
            cur.execute("DELETE FROM job_candidates WHERE job_id = ? AND candidate_id = ?", (job_id, candidate_id))
            
        elif action_type == "match_job":
            job_id = int(target_id)
            background_tasks.add_task(match_candidates_for_job, job_id)
            
        elif action_type == "reset_all":
            cur.execute("DELETE FROM candidate_metadata")
            try:
                if os.path.exists(CHROMA_PATH):
                    shutil.rmtree(CHROMA_PATH)
            except Exception:
                pass
                
        elif action_type == "approve_user":
            target_username = target_id
            cur.execute("UPDATE users SET is_approved = 1 WHERE LOWER(username) = LOWER(?)", (target_username,))
                
        cur.execute("UPDATE change_requests SET status = 'approved' WHERE id = ?", (request_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=500, detail=f"Failed to execute request: {str(e)}")
        
    conn.close()
    return {"status": "approved"}

@app.post("/api/admin/requests/{request_id}/reject")
def reject_change_request(request_id: int, request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT status, action_type, target_id FROM change_requests WHERE id = ?", (request_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Request not found")
    status, action_type, target_id = row
    if status != "pending":
        conn.close()
        raise HTTPException(status_code=400, detail="Request is already resolved")
        
    if action_type == "approve_user":
        cur.execute("DELETE FROM users WHERE LOWER(username) = LOWER(?)", (target_id,))
        
    cur.execute("UPDATE change_requests SET status = 'rejected' WHERE id = ?", (request_id,))
    conn.commit()
    conn.close()
    return {"status": "rejected"}

# ── Integrations Settings Endpoints ───────────────────────────────────────────
class IntegrationSettingsRequest(BaseModel):
    email_enabled: int
    imap_host: str
    imap_port: int
    smtp_host: str
    smtp_port: int
    email_user: str
    email_pass: str
    keywords: str
    drive_enabled: int

@app.get("/api/integrations")
def get_integrations_settings(request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT email_enabled, imap_host, imap_port, smtp_host, smtp_port, email_user, email_pass, keywords, drive_enabled FROM integrations_settings LIMIT 1")
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return {
            "email_enabled": 0, "imap_host": "imap.gmail.com", "imap_port": 993,
            "smtp_host": "smtp.gmail.com", "smtp_port": 587, "email_user": "",
            "email_pass": "", "keywords": "resume,alamaticz,solution,job", "drive_enabled": 0
        }
        
    masked_pass = ""
    if row[6]:
        masked_pass = "****"
        
    return {
        "email_enabled": row[0],
        "imap_host": row[1] or "imap.gmail.com",
        "imap_port": row[2] or 993,
        "smtp_host": row[3] or "smtp.gmail.com",
        "smtp_port": row[4] or 587,
        "email_user": row[5] or "",
        "email_pass": masked_pass,
        "keywords": row[7] or "resume,alamaticz,solution,job",
        "drive_enabled": row[8]
    }

@app.post("/api/integrations")
def save_integrations_settings(settings: IntegrationSettingsRequest, request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    
    cur.execute("SELECT id, email_pass FROM integrations_settings LIMIT 1")
    row = cur.fetchone()
    
    final_pass = settings.email_pass
    if final_pass == "****" and row:
        final_pass = row[1]
        
    if row:
        cur.execute("""
        UPDATE integrations_settings SET
            email_enabled = ?, imap_host = ?, imap_port = ?,
            smtp_host = ?, smtp_port = ?, email_user = ?,
            email_pass = ?, keywords = ?, drive_enabled = ?
        WHERE id = ?
        """, (settings.email_enabled, settings.imap_host, settings.imap_port,
              settings.smtp_host, settings.smtp_port, settings.email_user,
              final_pass, settings.keywords, settings.drive_enabled, row[0]))
    else:
        cur.execute("""
        INSERT INTO integrations_settings (
            email_enabled, imap_host, imap_port, smtp_host, smtp_port, email_user, email_pass, keywords, drive_enabled
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (settings.email_enabled, settings.imap_host, settings.imap_port,
              settings.smtp_host, settings.smtp_port, settings.email_user,
              final_pass, settings.keywords, settings.drive_enabled))
              
    if settings.email_enabled == 0:
        try:
            cur.execute("SELECT id, filename FROM candidate_metadata WHERE source = 'Import from Mail'")
            email_candidates = cur.fetchall()
            for cand_id, filename in email_candidates:
                if filename:
                    file_path = os.path.join(UPLOAD_DIR, filename)
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            print(f"Deleted email candidate file: {file_path}")
                        except Exception as file_err:
                            print(f"Error deleting file {file_path}: {file_err}")
                
                cur.execute("DELETE FROM job_candidates WHERE candidate_id = ?", (cand_id,))
                cur.execute("DELETE FROM candidate_metadata WHERE id = ?", (cand_id,))
            print(f"Deleted {len(email_candidates)} email-imported candidates because email sync was disabled.")
        except Exception as delete_err:
            print(f"Error purging email candidates: {delete_err}")

    conn.commit()
    conn.close()
    
    log_activity_db(username, "updated integrations settings")
    return {"status": "saved"}

@app.get("/api/integrations/status")
def test_integrations_connection(request: Request):
    username = request.headers.get("x-user-username")
    role = get_user_role(username)
    if role != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
        
    conn = sqlite3.connect(STATS_DB, timeout=30.0)
    cur = conn.cursor()
    cur.execute("SELECT email_enabled, imap_host, imap_port, email_user, email_pass FROM integrations_settings LIMIT 1")
    row = cur.fetchone()
    conn.close()
    
    if not row or not row[0]:
        return {"status": "disabled", "message": "Email integration is disabled."}
        
    imap_host = row[1]
    imap_port = row[2] or 993
    email_user = row[3]
    email_pass = row[4]
    
    if not imap_host or not email_user or not email_pass:
        return {"status": "unconfigured", "message": "Credentials are not fully configured."}
        
    import imaplib
    try:
        mail = imaplib.IMAP4_SSL(imap_host, imap_port, timeout=10)
        mail.login(email_user, email_pass)
        mail.logout()
        return {"status": "connected", "message": "Successfully connected to Gmail!"}
    except Exception as e:
        err_msg = str(e)
        if "Application-specific password required" in err_msg:
            return {"status": "error", "message": "Authentication failed: Application-specific password required."}
        return {"status": "error", "message": f"Connection failed: {err_msg}"}

def poll_emails_and_process():
    import time
    import imaplib
    import email
    from email.header import decode_header
    import hashlib
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    print("INFO: Starting background email worker thread...")
    while True:
        try:
            # Fetch settings from DB
            conn = sqlite3.connect(STATS_DB, timeout=30.0)
            cur = conn.cursor()
            cur.execute("""
                SELECT email_enabled, imap_host, imap_port, smtp_host, smtp_port, 
                       email_user, email_pass, keywords 
                FROM integrations_settings LIMIT 1
            """)
            row = cur.fetchone()
            conn.close()

            if not row or not row[0]: # Not enabled or not configured
                time.sleep(15)
                continue

            email_enabled, imap_host, imap_port, smtp_host, smtp_port, email_user, email_pass, keywords_str = row
            if not imap_host or not email_user or not email_pass:
                time.sleep(15)
                continue

            # Parse keywords
            keywords = [k.strip().lower() for k in keywords_str.split(",") if k.strip()]
            if not keywords:
                keywords = ["resume", "alamaticz", "solution", "job"]

            # Connect IMAP
            mail = imaplib.IMAP4_SSL(imap_host, imap_port or 993, timeout=15)
            mail.login(email_user, email_pass)
            mail.select("inbox")

            # Search unseen or recent messages to catch already read messages
            unseen_nums = []
            status_unseen, response_unseen = mail.search(None, "UNSEEN")
            if status_unseen == "OK" and response_unseen[0]:
                unseen_nums = response_unseen[0].split()

            all_nums = []
            status_all, response_all = mail.search(None, "ALL")
            if status_all == "OK" and response_all[0]:
                all_nums = response_all[0].split()

            # Take the last 30 messages in the inbox (regardless of read status)
            recent_nums = all_nums[-30:]

            # Combine them, deduplicate, and sort as integers to process in chronological order
            msg_nums_set = set(unseen_nums + recent_nums)
            msg_nums = sorted(list(msg_nums_set), key=lambda x: int(x))
            for num in msg_nums:
                try:
                    # Fetch message-id and headers without marking as seen
                    status, header_data = mail.fetch(num, '(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID)])')
                    msg_id = None
                    if status == "OK" and header_data[0]:
                        header_text = header_data[0][1].decode('utf-8', errors='ignore')
                        match = re.search(r'Message-ID:\s*<([^>]+)>', header_text, re.IGNORECASE)
                        if match:
                            msg_id = match.group(1)
                    
                    if not msg_id:
                        # Fallback: hash of headers to make a pseudo ID
                        status, header_data2 = mail.fetch(num, '(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE FROM)])')
                        header_text2 = ""
                        if status == "OK" and header_data2[0]:
                            header_text2 = header_data2[0][1].decode('utf-8', errors='ignore')
                        msg_id = hashlib.md5(header_text2.encode('utf-8', errors='ignore')).hexdigest()

                    # Check if already processed
                    conn = sqlite3.connect(STATS_DB, timeout=30.0)
                    cur = conn.cursor()
                    cur.execute("SELECT 1 FROM processed_emails WHERE msg_uid = ?", (msg_id,))
                    exists = cur.fetchone()
                    conn.close()

                    if exists:
                        continue

                    # Fetch full message content without marking read
                    status, msg_data = mail.fetch(num, '(BODY.PEEK[])')
                    if status != "OK" or not msg_data[0]:
                        continue

                    raw_email = msg_data[0][1]
                    msg = email.message_from_bytes(raw_email)

                    # Extract Subject, From, Body
                    subject_header = msg.get("Subject", "")
                    decoded_subject = ""
                    for part, encoding in decode_header(subject_header):
                        if isinstance(part, bytes):
                            decoded_subject += part.decode(encoding or "utf-8", errors="ignore")
                        else:
                            decoded_subject += str(part)
                    
                    from_header = msg.get("From", "")
                    decoded_from = ""
                    for part, encoding in decode_header(from_header):
                        if isinstance(part, bytes):
                            decoded_from += part.decode(encoding or "utf-8", errors="ignore")
                        else:
                            decoded_from += str(part)

                    # Extract plain text body
                    body_text = ""
                    attachments = []
                    if msg.is_multipart():
                        for part in msg.walk():
                            content_type = part.get_content_type()
                            content_disposition = str(part.get("Content-Disposition"))
                            if content_type == "text/plain" and "attachment" not in content_disposition:
                                payload = part.get_payload(decode=True)
                                if payload:
                                    body_text += payload.decode(part.get_content_charset() or "utf-8", errors="ignore")
                            elif "attachment" in content_disposition or part.get_filename():
                                filename = part.get_filename()
                                if filename:
                                    decoded_filename = ""
                                    for filename_part, encoding in decode_header(filename):
                                        if isinstance(filename_part, bytes):
                                            decoded_filename += filename_part.decode(encoding or "utf-8", errors="ignore")
                                        else:
                                            decoded_filename += str(filename_part)
                                    if decoded_filename:
                                        attachments.append((decoded_filename, part.get_payload(decode=True)))
                    else:
                        payload = msg.get_payload(decode=True)
                        if payload:
                            body_text = payload.decode(msg.get_content_charset() or "utf-8", errors="ignore")
                    # Check if this email is from an existing candidate (via Ref tag in Subject, or matching sender email)
                    import email.utils
                    sender_name, sender_email = email.utils.parseaddr(decoded_from)
                    matched_candidate_row = None

                    # 1. Search by reference ID in subject
                    match_ref = re.search(r'Ref:\s*CAND-(\d+)', decoded_subject, re.IGNORECASE)
                    if match_ref:
                        try:
                            ref_candidate_id = int(match_ref.group(1))
                            conn = sqlite3.connect(STATS_DB, timeout=30.0)
                            conn.row_factory = sqlite3.Row
                            cur = conn.cursor()
                            cur.execute("SELECT * FROM candidate_metadata WHERE id = ?", (ref_candidate_id,))
                            matched_candidate_row = cur.fetchone()
                            conn.close()
                        except Exception as e_ref:
                            print(f"ERROR matching by Ref ID in subject: {e_ref}")
                    # 2. Fall back to search by sender email address
                    if not matched_candidate_row and sender_email:
                        try:
                            conn = sqlite3.connect(STATS_DB, timeout=30.0)
                            conn.row_factory = sqlite3.Row
                            cur = conn.cursor()
                            cur.execute("SELECT * FROM candidate_metadata WHERE LOWER(sender_email) = ? OR LOWER(email) = ? ORDER BY id DESC LIMIT 1", (sender_email.lower(), sender_email.lower()))
                            matched_candidate_row = cur.fetchone()
                            conn.close()
                        except Exception as e_email:
                            print(f"ERROR matching by sender email: {e_email}")

                    if matched_candidate_row:
                        existing_candidate = dict(matched_candidate_row)
                        print(f"INFO: Identified follow-up email from existing candidate ID {existing_candidate['id']} ({existing_candidate['full_name']})")
                        
                        # Check if a new resume file is attached and download/index it
                        new_filename = None
                        resume_text = ""
                        if attachments:
                            has_resume = any(
                                fname.lower().endswith((".pdf", ".docx"))
                                for fname, _ in attachments if fname
                            )
                            if has_resume:
                                for fname, content in attachments:
                                    if not content:
                                        continue
                                    f_lower = fname.lower()
                                    if f_lower.endswith(".pdf") or f_lower.endswith(".docx"):
                                        safe_name = re.sub(r'[^a-zA-Z0-9._-]', '_', fname)
                                        safe_name = f"mail_{hashlib.md5(msg_id.encode()).hexdigest()[:8]}_{safe_name}"
                                        fpath = os.path.join(UPLOAD_DIR, safe_name)
                                        with open(fpath, "wb") as f_out:
                                            f_out.write(content)
                                        new_filename = safe_name
                                        
                                        # Index in ChromaDB and extract text
                                        try:
                                            if safe_name.lower().endswith(".pdf"):
                                                loader = SafePyMuPDFLoader(fpath)
                                            else:
                                                loader = Docx2txtLoader(fpath)
                                            docs = loader.load()
                                            resume_text = "\n".join([d.page_content for d in docs])
                                            embeddings, _ = get_models()
                                            for d in docs:
                                                d.metadata['source'] = safe_name
                                            splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
                                            chunks = splitter.split_documents(docs)
                                            Chroma.from_documents(chunks, embeddings, persist_directory=CHROMA_PATH)
                                        except Exception as chroma_err:
                                            print(f"Error adding updated resume to Chroma: {chroma_err}")
                                        break

                        # Use EXTRACT_PROMPT to parse the combined new resume content and/or email text
                        combined_text = resume_text[:7000] if resume_text else ""
                        if body_text:
                            combined_text += f"\n\n=== EMAIL MESSAGE BODY ===\n{body_text}\n=========================="

                        conn = sqlite3.connect(STATS_DB, timeout=30.0)
                        cur = conn.cursor()
                        cur.execute("SELECT col_key, col_label, description FROM custom_columns")
                        custom_cols = cur.fetchall()
                        conn.close()
                        
                        custom_fields_str = ""
                        if custom_cols:
                            for col_key, col_label, desc in custom_cols:
                                desc_str = f"Extract data corresponding to column heading '{col_label}'"
                                if desc:
                                    desc_str += f" ({desc})"
                                custom_fields_str += f',\n  "{col_key}": "<{desc_str}>"'
                                
                        prompt_str = EXTRACT_PROMPT.format(text=combined_text, custom_fields=custom_fields_str)
                        
                        _, llm = get_models()
                        parsed_data = {}
                        try:
                            # Add a simple retry mechanism for rate limits
                            max_retries = 3
                            resp = None
                            for attempt in range(max_retries):
                                try:
                                    resp = llm.invoke([HumanMessage(content=prompt_str)])
                                    break
                                except Exception as api_err:
                                    if "429" in str(api_err) and attempt < max_retries - 1:
                                        import time
                                        time.sleep(3)
                                        continue
                                    raise api_err
                                    
                            if resp is not None:
                                raw_resp = resp.content.strip()
                                if "```json" in raw_resp:
                                    raw_resp = raw_resp.split("```json")[1].split("```")[0].strip()
                                elif "```" in raw_resp:
                                    raw_resp = raw_resp.split("```")[1].split("```")[0].strip()
                                start_idx, end_idx = raw_resp.find('{'), raw_resp.rfind('}')
                                if start_idx != -1 and end_idx != -1:
                                    raw_resp = raw_resp[start_idx:end_idx+1]
                                parsed_data = json.loads(raw_resp)
                        except Exception as llm_err:
                            print(f"ERROR parsing follow-up email/resume via LLM: {llm_err}")

                        # Normalization & Validation
                        # Phone: Keep only digits and +
                        if 'phone' in parsed_data and parsed_data['phone']:
                            parsed_data['phone'] = re.sub(r'[^\d+]', '', str(parsed_data['phone']))
                            
                        # Email: Basic validation
                        if 'email' in parsed_data and parsed_data['email']:
                            if '@' not in str(parsed_data['email']):
                                parsed_data['email'] = ""
                                
                        # Experience: Force float
                        for exp_field in ['total_experience', 'pega_experience', 'cdh_exp']:
                            if exp_field in parsed_data and parsed_data[exp_field] not in [None, ""]:
                                try:
                                    match = re.search(r'\d+(\.\d+)?', str(parsed_data[exp_field]))
                                    parsed_data[exp_field] = float(match.group()) if match else 0.0
                                except Exception:
                                    parsed_data[exp_field] = 0.0
                                    
                        # Integer fields
                        for num_field in ['notice_period', 'availability_in_days']:
                            if num_field in parsed_data and parsed_data[num_field] not in [None, ""]:
                                np_str = str(parsed_data[num_field]).lower()
                                if 'immediate' in np_str:
                                    parsed_data[num_field] = 0
                                else:
                                    try:
                                        match = re.search(r'\d+', np_str)
                                        val = int(match.group()) if match else ""
                                        if match and 'month' in np_str:
                                            val = val * 30
                                        parsed_data[num_field] = val
                                    except Exception:
                                        parsed_data[num_field] = ""

                        # Update existing row with all dynamic columns
                        updates = {}
                        conn = sqlite3.connect(STATS_DB, timeout=30.0)
                        cur = conn.cursor()
                        cur.execute("PRAGMA table_info(candidate_metadata)")
                        allowed_cols = {c[1] for c in cur.fetchall()}
                        conn.close()

                        for k, v in parsed_data.items():
                            if k in allowed_cols and k != 'id' and k != 'filename':
                                if v is not None and str(v).strip() not in ("", "null", "None"):
                                    updates[k] = v

                        if new_filename:
                            updates['filename'] = new_filename
                        if sender_email:
                            updates['sender_email'] = sender_email

                        # Merge email message body
                        old_msg = existing_candidate.get('email_message') or ""
                        new_msg = f"{old_msg}\n\n=== FOLLOW-UP EMAIL MESSAGE ===\n{body_text}".strip()
                        updates['email_message'] = new_msg
                        
                        # Reset formatted resume cache
                        updates['formatted_json'] = None

                        if updates:
                            try:
                                conn = sqlite3.connect(STATS_DB, timeout=30.0)
                                cur = conn.cursor()
                                set_clause = ", ".join(f"{col}=?" for col in updates)
                                cur.execute(f"UPDATE candidate_metadata SET {set_clause} WHERE id=?", list(updates.values()) + [existing_candidate['id']])
                                conn.commit()
                                conn.close()
                                print(f"INFO: Successfully updated candidate ID {existing_candidate['id']} with follow-up details.")
                            except Exception as db_update_err:
                                print(f"ERROR updating candidate from follow-up email: {db_update_err}")

                        # Recheck missing fields on updated profile
                        try:
                            conn = sqlite3.connect(STATS_DB, timeout=30.0)
                            conn.row_factory = sqlite3.Row
                            cur = conn.cursor()
                            cur.execute("SELECT * FROM candidate_metadata WHERE id=?", (existing_candidate['id'],))
                            updated_candidate = dict(cur.fetchone())
                            conn.close()

                            candidate_name = updated_candidate.get('full_name') or sender_name or 'Candidate'
                            missing_fields = []
                            total_exp = updated_candidate.get('total_experience')
                            if total_exp is None or str(total_exp).strip() == "" or float(total_exp) == 0.0:
                                missing_fields.append("Total years of experience")
                            
                            pega_exp = updated_candidate.get('pega_experience')
                            cdh_exp = updated_candidate.get('cdh_exp')
                            has_pega = pega_exp is not None and str(pega_exp).strip() != "" and float(pega_exp) > 0.0
                            has_cdh = cdh_exp is not None and str(cdh_exp).strip() != "" and float(cdh_exp) > 0.0
                            if not has_pega and not has_cdh:
                                missing_fields.append("Relevant experience for this role")
                            
                            ctc = updated_candidate.get('ctc')
                            if not ctc or str(ctc).strip() in ("", "—", "-", "None", "null"):
                                missing_fields.append("Current CTC")
                            
                            expected_ctc = updated_candidate.get('expected_ctc')
                            if not expected_ctc or str(expected_ctc).strip() in ("", "—", "-", "None", "null"):
                                missing_fields.append("Expected CTC")
                            
                            notice_period = updated_candidate.get('notice_period')
                            if notice_period is None or str(notice_period).strip() in ("", "—", "-", "None", "null"):
                                missing_fields.append("Notice period / Earliest joining date")
                            
                            current_location = updated_candidate.get('current_location')
                            if not current_location or str(current_location).strip() in ("", "—", "-", "None", "null"):
                                missing_fields.append("Current location")
                            
                            pref_locations = updated_candidate.get('pref_locations')
                            if not pref_locations or str(pref_locations).strip() in ("", "—", "-", "None", "null"):
                                missing_fields.append("Preferred work location(s)")
                            
                            linkedin = updated_candidate.get('linkedin')
                            if not linkedin or str(linkedin).strip() in ("", "—", "-", "None", "null"):
                                missing_fields.append("LinkedIn profile URL")

                            if missing_fields:
                                missing_list_str = "\n".join(f"* {field}:" for field in missing_fields)
                                body_reply = f"Dear {candidate_name},\n\n" \
                                             f"Thank you for providing the details.\n\n" \
                                             f"We noticed that the following required details are still missing. Kindly share them to proceed further:\n\n" \
                                             f"{missing_list_str}\n\n" \
                                             f"Once we receive the above information, our recruitment team will review your profile and get back to you regarding the next steps in the selection process.\n\n" \
                                             f"We look forward to hearing from you.\n\n" \
                                             f"Best regards,\n\n" \
                                             f"HR Team\n" \
                                             f"Alamaticz Solutions"
                            else:
                                body_reply = f"Dear {candidate_name},\n\n" \
                                             f"Thank you for your interest in Alamaticz Solutions and for submitting your application.\n\n" \
                                             f"We have successfully received all required details. Our recruitment team will review your profile and get back to you regarding the next steps in the selection process.\n\n" \
                                             f"Best regards,\n\n" \
                                             f"HR Team\n" \
                                             f"Alamaticz Solutions"

                            ack_msg = MIMEMultipart()
                            ack_msg['From'] = email_user
                            ack_msg['To'] = sender_email
                            ack_msg['Subject'] = f"Re: {decoded_subject} (Ref: CAND-{updated_candidate['id']})"
                            ack_msg.attach(MIMEText(body_reply, 'plain'))

                            s_server = smtplib.SMTP(smtp_host, smtp_port or 587, timeout=10.0)
                            s_server.starttls()
                            s_server.login(email_user, email_pass)
                            s_server.sendmail(email_user, sender_email, ack_msg.as_string())
                            s_server.quit()
                            print(f"INFO: Sent follow-up acknowledgment to {sender_email}")
                        except Exception as reply_err:
                            print(f"ERROR sending follow-up email reply: {reply_err}")

                        # Mark email as read and processed
                        try:
                            mail.store(num, '+FLAGS', '\\Seen')
                            conn = sqlite3.connect(STATS_DB, timeout=30.0)
                            cur = conn.cursor()
                            cur.execute("INSERT INTO processed_emails (msg_uid) VALUES (?)", (msg_id,))
                            conn.commit()
                            conn.close()
                        except Exception as mark_err:
                            print(f"ERROR marking email as processed: {mark_err}")
                            
                        continue

                    # Check if the email contains any resume attachments (.pdf or .docx)
                    has_resume_attachment = False
                    if attachments:
                        has_resume_attachment = any(
                            fname.lower().endswith((".pdf", ".docx"))
                            for fname, _ in attachments if fname
                        )

                    # Always match if there is a resume attachment, otherwise fall back to keyword match
                    match_found = has_resume_attachment
                    if not match_found:
                        search_content = f"{decoded_subject} {body_text}".lower()
                        for kw in keywords:
                            if kw in search_content:
                                match_found = True
                                break

                    processed_attachment = False
                    if match_found and attachments:
                        # Only proceed if there is at least one resume attachment (.pdf or .docx)
                        has_resume = any(
                            fname.lower().endswith(".pdf") or fname.lower().endswith(".docx")
                            for fname, _ in attachments if fname
                        )
                        if has_resume:
                            # Process attachments
                            for fname, content in attachments:
                                if not content:
                                    continue
                                f_lower = fname.lower()
                                if f_lower.endswith(".pdf") or f_lower.endswith(".docx"):
                                    safe_name = re.sub(r'[^a-zA-Z0-9._-]', '_', fname)
                                    safe_name = f"mail_{hashlib.md5(msg_id.encode()).hexdigest()[:8]}_{safe_name}"
                                    fpath = os.path.join(UPLOAD_DIR, safe_name)
                                    with open(fpath, "wb") as f_out:
                                        f_out.write(content)
                                    
                                    try:
                                        process_resume(safe_name, fpath, is_approved=1, username="email_worker", email_message=body_text, sender_email=sender_email)
                                        processed_attachment = True
                                    except Exception as e_proc:
                                        print(f"ERROR: Failed processing resume {safe_name} from email: {e_proc}")
                        
                        if processed_attachment:
                            # Mark email as read on server
                            mail.store(num, '+FLAGS', '\\Seen')
                            
                            # Send auto acknowledgment
                            try:
                                import email.utils
                                sender_name, sender_email = email.utils.parseaddr(from_header)
                                if sender_email:
                                    # Fetch parsed candidate metadata to find missing fields
                                    candidate = None
                                    try:
                                        conn = sqlite3.connect(STATS_DB, timeout=30.0)
                                        conn.row_factory = sqlite3.Row
                                        cur = conn.cursor()
                                        cur.execute("SELECT * FROM candidate_metadata WHERE filename = ? ORDER BY id DESC LIMIT 1", (safe_name,))
                                        candidate_row = cur.fetchone()
                                        if candidate_row:
                                            candidate = dict(candidate_row)
                                        conn.close()
                                    except Exception as db_err:
                                        print(f"ERROR: Failed to fetch candidate metadata for auto-acknowledgement: {db_err}")

                                    candidate_name = 'Candidate'
                                    missing_fields = []
                                    if candidate:
                                        candidate_name = candidate.get('full_name') or sender_name or 'Candidate'
                                        
                                        # 1. Total years of experience
                                        total_exp = candidate.get('total_experience')
                                        if total_exp is None or str(total_exp).strip() == "" or float(total_exp) == 0.0:
                                            missing_fields.append("Total years of experience")
                                            
                                        # 2. Relevant experience for this role
                                        pega_exp = candidate.get('pega_experience')
                                        cdh_exp = candidate.get('cdh_exp')
                                        has_pega = pega_exp is not None and str(pega_exp).strip() != "" and float(pega_exp) > 0.0
                                        has_cdh = cdh_exp is not None and str(cdh_exp).strip() != "" and float(cdh_exp) > 0.0
                                        if not has_pega and not has_cdh:
                                            missing_fields.append("Relevant experience for this role")
                                            
                                        # 3. Current CTC
                                        ctc = candidate.get('ctc')
                                        if not ctc or str(ctc).strip() in ("", "—", "-", "None", "null"):
                                            missing_fields.append("Current CTC")
                                            
                                        # 4. Expected CTC
                                        expected_ctc = candidate.get('expected_ctc')
                                        if not expected_ctc or str(expected_ctc).strip() in ("", "—", "-", "None", "null"):
                                            missing_fields.append("Expected CTC")
                                            
                                        # 5. Notice period / Earliest joining date
                                        notice_period = candidate.get('notice_period')
                                        if notice_period is None or str(notice_period).strip() in ("", "—", "-", "None", "null"):
                                            missing_fields.append("Notice period / Earliest joining date")
                                            
                                        # 6. Current location
                                        current_location = candidate.get('current_location')
                                        if not current_location or str(current_location).strip() in ("", "—", "-", "None", "null"):
                                            missing_fields.append("Current location")
                                            
                                        # 7. Preferred work location(s)
                                        pref_locations = candidate.get('pref_locations')
                                        if not pref_locations or str(pref_locations).strip() in ("", "—", "-", "None", "null"):
                                            missing_fields.append("Preferred work location(s)")
                                            
                                        # 8. LinkedIn profile URL
                                        linkedin = candidate.get('linkedin')
                                        if not linkedin or str(linkedin).strip() in ("", "—", "-", "None", "null"):
                                            missing_fields.append("LinkedIn profile URL")
                                    else:
                                        candidate_name = sender_name or 'Candidate'
                                        missing_fields = [
                                            "Total years of experience",
                                            "Relevant experience for this role",
                                            "Current CTC",
                                            "Expected CTC",
                                            "Notice period / Earliest joining date",
                                            "Current location",
                                            "Preferred work location(s)",
                                            "LinkedIn profile URL"
                                        ]

                                    if missing_fields:
                                        missing_list_str = "\n".join(f"* {field}:" for field in missing_fields)
                                        body_reply = f"Dear {candidate_name},\n\n" \
                                                     f"Thank you for your interest in Alamaticz Solutions and for submitting your application.\n\n" \
                                                     f"We appreciate the time you have taken to apply for this opportunity. To help us evaluate your profile further, kindly share the following details:\n\n" \
                                                     f"{missing_list_str}\n\n" \
                                                     f"Once we receive the above information, our recruitment team will review your profile and get back to you regarding the next steps in the selection process.\n\n" \
                                                     f"We look forward to hearing from you.\n\n" \
                                                     f"Best regards,\n\n" \
                                                     f"HR Team\n" \
                                                     f"Alamaticz Solutions"
                                    else:
                                        body_reply = f"Dear {candidate_name},\n\n" \
                                                     f"Thank you for your interest in Alamaticz Solutions and for submitting your application.\n\n" \
                                                     f"We appreciate the time you have taken to apply for this opportunity. Our recruitment team will review your profile and get back to you regarding the next steps in the selection process.\n\n" \
                                                     f"Best regards,\n\n" \
                                                     f"HR Team\n" \
                                                     f"Alamaticz Solutions"

                                    ack_msg = MIMEMultipart()
                                    ack_msg['From'] = email_user
                                    ack_msg['To'] = sender_email
                                    cand_id = candidate.get('id') if candidate else 0
                                    if cand_id:
                                        ack_msg['Subject'] = f"Re: {decoded_subject} (Ref: CAND-{cand_id})"
                                    else:
                                        ack_msg['Subject'] = f"Re: {decoded_subject}"
                                    ack_msg.attach(MIMEText(body_reply, 'plain'))
                                    
                                    s_server = smtplib.SMTP(smtp_host, smtp_port or 587, timeout=10.0)
                                    s_server.starttls()
                                    s_server.login(email_user, email_pass)
                                    s_server.sendmail(email_user, sender_email, ack_msg.as_string())
                                    s_server.quit()
                                    print(f"INFO: Sent auto-acknowledgement to {sender_email}")
                            except Exception as smtp_err:
                                print(f"ERROR: Failed sending auto-acknowledgement: {smtp_err}")

                    # Insert to processed_emails
                    conn = sqlite3.connect(STATS_DB, timeout=30.0)
                    cur = conn.cursor()
                    cur.execute("INSERT OR IGNORE INTO processed_emails (msg_uid) VALUES (?)", (msg_id,))
                    conn.commit()
                    conn.close()

                except Exception as msg_err:
                    print(f"ERROR: Failed processing single email: {msg_err}")

            mail.logout()

        except Exception as conn_err:
            print(f"ERROR: Email background loop connection error: {conn_err}")
        
        time.sleep(30)

# ── Serve React Frontend ───────────────────────────────────────────────────────
FRONTEND_DIST = os.path.join(PROJECT_ROOT, "frontend", "dist")
if os.path.exists(FRONTEND_DIST):
    app.mount("/", StaticFiles(directory=FRONTEND_DIST, html=True), name="frontend")

@app.exception_handler(StarletteHTTPException)
async def catch_all_spa_routes(request: Request, exc: StarletteHTTPException):
    # If a 404 occurs and it's not an API call, serve the React index.html for client-side routing
    if exc.status_code == 404 and not request.url.path.startswith("/api/"):
        index_file = os.path.join(FRONTEND_DIST, "index.html")
        if os.path.exists(index_file):
            return FileResponse(index_file)
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})

